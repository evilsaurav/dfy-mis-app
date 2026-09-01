import openpyxl

wb = openpyxl.load_workbook("templates/template_Jamui.xlsx", data_only=False)

print("Sheet names:", wb.sheetnames[:10])

# Check a daily sheet, e.g., '1st' or '1ST' or ordinal
for name in wb.sheetnames:
    if name.lower().startswith("1"):
        ws = wb[name]
        print(f"\n--- Inspecting Daily Sheet: '{name}' ---")
        print(f"Max rows: {ws.max_row}, Max cols: {ws.max_column}")
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(25, ws.max_column + 1))]
            # Print non-empty
            if any(row_vals):
                print(f"Row {r:2d}: {row_vals[:10]}")
        break

# Check CONSOLIDATED SHEET
if "CONSOLIDATED SHEET" in wb.sheetnames:
    ws_cons = wb["CONSOLIDATED SHEET"]
    print(f"\n--- Inspecting CONSOLIDATED SHEET ---")
    print(f"Max rows: {ws_cons.max_row}, Max cols: {ws_cons.max_column}")
    for r in range(1, min(15, ws_cons.max_row + 1)):
        row_vals = [ws_cons.cell(row=r, column=c).value for c in range(1, min(20, ws_cons.max_column + 1))]
        print(f"Row {r:2d}: {row_vals[:12]}")
