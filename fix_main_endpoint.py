from fastapi.responses import StreamingResponse
import io
import openpyxl
import re

def safe_filename(district: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", district.strip())
    return cleaned.strip("_") or "UNKNOWN"

def ordinal(n: int) -> str:
    if n == 1:
        return "1ST"
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

@app.get("/download-kpi-workbook")
async def download_kpi_workbook(district: str):
    try:
        # Load the static template
        safe_dist = safe_filename(district)
        template_path = f"templates/template_{safe_dist}.xlsx"
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail=f"Template for {district} not found on server.")
        
        wb = openpyxl.load_workbook(template_path)
        
        # Fetch reports from Firestore
        docs = db.collection("daily_field_reports").where("working_place", "==", district).stream()
        
        # KPI Map
        KPI_MAP = {
            "notification_ids": 3, "hiv_dm_ids": 4, "dbt_ids": 5, "sample_collection_ids": 6,
            "sample_tested_ids": 7, "outcome_assigned_ids": 8, "home_visit_ids": 9,
            "contact_tracing_ids": 10, "follow_up_ids": 11, "face_to_face_ids": 12,
            "presumptive_ids": 13, "documents_ids": 14, "fdc_provided_ids": 15, "kit_consumption_ids": 16,
            "differentiated_tb_ids": 17, "tpt_treatment_start_ids": 18, "tpt_presumptive_ids": 19,
            "adhar_face_authentication_ids": 20, "consent_with_id_ids": 21
        }
        
        for doc in docs:
            data = doc.to_dict()
            date_str = data.get("date_of_reporting")
            if not date_str:
                continue
                
            try:
                day_int = int(date_str.split('-')[2])
            except:
                continue
                
            tab_name = ordinal(day_int)
            fo_name = data.get("fo_name", "").strip().lower()
            
            # 1. Update Daily Tab
            if tab_name in wb.sheetnames:
                ws = wb[tab_name]
                start_row = -1
                for r_idx in range(1, ws.max_row + 1):
                    cell_val = ws.cell(row=r_idx, column=1).value
                    if cell_val and str(cell_val).strip().lower() == fo_name:
                        start_row = r_idx
                        break
                
                if start_row != -1:
                    BLOCK_SIZE = 40
                    for key, col_idx in KPI_MAP.items():
                        ids = data.get(key) or []
                        for i in range(BLOCK_SIZE):
                            cell = ws.cell(row=start_row + i, column=col_idx)
                            if i < len(ids):
                                cell.value = str(ids[i])
                            # Do not need to blank out because it's a fresh template anyway!
                            # Wait, yes! It's a blank template loaded from disk, so no old ghosts.

            # 2. Update CONSOLIDATED SHEET (Right Section)
            if "CONSOLIDATED SHEET" in wb.sheetnames:
                ws_cons = wb["CONSOLIDATED SHEET"]
                base_col = -1
                for c_idx in range(1, ws_cons.max_column + 1):
                    cell_val = ws.cell(row=1, column=c_idx).value
                    if cell_val and str(cell_val).strip().lower() == fo_name:
                        base_col = c_idx
                        break
                
                if base_col != -1:
                    cons_row = 2 + day_int
                    ws_cons.cell(row=cons_row, column=base_col + 1, value="Present")
                    ws_cons.cell(row=cons_row, column=base_col + 2, value=len(data.get("hiv_dm_ids") or []))
                    ws_cons.cell(row=cons_row, column=base_col + 3, value=len(data.get("dbt_ids") or []))
                    
        # Save to memory and return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="KPI_Report_{safe_dist}.xlsx"'
        }
        return StreamingResponse(
            output, 
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"Error generating KPI workbook: {e}")
        raise HTTPException(status_code=500, detail=str(e))
