
# Professional OpenPyXL Border & Style Helpers
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

EXCEL_HEADER_BORDER = Border(
    left=Side(style='thin', color='94A3B8'),
    right=Side(style='thin', color='94A3B8'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='medium', color='1E293B')
)

EXCEL_TOTAL_ROW_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='double', color='1E293B')
)

EXCEL_CLUSTER_DIVIDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='medium', color='64748B'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

def style_excel_worksheet(ws, header_fill_color="4F46E5"):
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = EXCEL_HEADER_BORDER
        
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = EXCEL_THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

import zipfile
import firebase_admin
from firebase_admin import credentials, firestore, storage
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import pandas as pd
import io
import os
import json
import uuid

firebase_creds_env = os.environ.get("FIREBASE_CREDENTIALS")
if firebase_creds_env:
    cred_dict = json.loads(firebase_creds_env)
    cred = credentials.Certificate(cred_dict)
else:
    cred = credentials.Certificate("firebase_key.json")

if not firebase_admin._apps:
    firebase_admin.initialize_app(cred, {
        'storageBucket': 'dfy-reporting-mis.appspot.com'
    })
db = firestore.client()

import time
import asyncio
from typing import Dict, Any, Tuple

class SimpleTTLCache:
    def __init__(self, default_ttl: int = 30):
        self._cache: Dict[str, Tuple[float, Any]] = {}
        self.default_ttl = default_ttl

    def get(self, key: str):
        if key in self._cache:
            exp, val = self._cache[key]
            if time.time() < exp:
                return val
            else:
                del self._cache[key]
        return None

    def set(self, key: str, val: Any, ttl: Optional[int] = None):
        t = ttl if ttl is not None else self.default_ttl
        self._cache[key] = (time.time() + t, val)

    def delete(self, key: str):
        if key in self._cache:
            del self._cache[key]

    def delete_prefix(self, prefix: str):
        keys_to_del = [k for k in self._cache if k.startswith(prefix)]
        for k in keys_to_del:
            del self._cache[k]

    def clear(self):
        self._cache.clear()

cache = SimpleTTLCache(default_ttl=30)


app = FastAPI(title="DFY Daily Activity API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class PinCheck(BaseModel):
    working_place: str
    fo_name: str
    pin: str

class DailyActivityReport(BaseModel):
    date_of_reporting: Optional[str] = None
    working_place: str
    fo_name: str
    pin: str
    
    notification_ids: List[str] = []
    hiv_dm_ids: List[str] = []
    dbt_ids: List[str] = []
    sample_collection_ids: List[str] = []
    sample_tested_ids: List[str] = []
    outcome_assigned_ids: List[str] = []
    home_visit_ids: List[str] = []
    contact_tracing_ids: List[str] = []
    follow_up_ids: List[str] = []
    face_to_face_ids: List[str] = []
    presumptive_ids: List[str] = []
    documents_ids: List[str] = []
    fdc_provided_ids: List[str] = []
    kit_consumption_ids: List[str] = []
    differentiated_tb_ids: List[str] = []
    tpt_treatment_start_ids: List[str] = []
    tpt_presumptive_ids: List[str] = []
    adhar_face_authentication_ids: List[str] = []
    consent_with_id_ids: List[str] = []
    culture_dst_ids: List[str] = []
    
    remark: Optional[str] = ""
    
    doctor_store_visits_count: Optional[int] = 0
    visited_names: List[str] = []
    morning_km: Optional[int] = 0
    evening_km: Optional[int] = 0
    morning_km_photo_url: Optional[str] = ""
    evening_km_photo_url: Optional[str] = ""
    is_override_used: Optional[bool] = False

class DashboardRequest(BaseModel):
    month_prefix: str
    districts: Optional[str] = None

@app.post("/admin/dashboard-data")
async def get_dashboard_data(req: DashboardRequest):
    try:
        start_date = f"{req.month_prefix}-01"
        end_date = f"{req.month_prefix}-31"
        
        allowed_dist_set = None
        if req.districts and req.districts.strip() and req.districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in req.districts.split(",") if d.strip()])

        docs = db.collection("daily_field_reports")\
            .where("date_of_reporting", ">=", start_date)\
            .where("date_of_reporting", "<=", end_date)\
            .stream()
            
        records = []
        for doc in docs:
            data = doc.to_dict()
            wp = data.get("working_place", "Unknown")
            if allowed_dist_set and wp not in allowed_dist_set:
                continue

            records.append({
                "date": data.get("date_of_reporting", ""),
                "working_place": wp,
                "fo_name": data.get("fo_name", "Unknown"),
                
                # Big 5
                "total_km": data.get("total_km", 0) or 0,
                "notifications": len(data.get("notification_ids", [])),
                "tests": len(data.get("sample_tested_ids", [])),
                "presumptive": len(data.get("presumptive_ids", [])),
                "doctor_visits": len(data.get("visited_names", [])),
                
                # Group 1
                "hiv_dm": len(data.get("hiv_dm_ids", [])),
                "dbt": len(data.get("dbt_ids", [])),
                
                # Group 2
                "sample_collection": len(data.get("sample_collection_ids", [])),
                "outcome_assigned": len(data.get("outcome_assigned_ids", [])),
                
                # Group 3
                "home_visits": len(data.get("home_visit_ids", [])),
                "contact_tracing": len(data.get("contact_tracing_ids", [])),
                "follow_ups": len(data.get("follow_up_ids", [])),
                "face_to_face": len(data.get("face_to_face_ids", [])),
                
                # Group 4
                "documents": len(data.get("documents_ids", [])),
                "fdc_provided": len(data.get("fdc_provided_ids", [])),
                "kit_consumption": len(data.get("kit_consumption_ids", [])),
                
                # Group 5 (New Fields & Special)
                "differentiated_tb": len(data.get("differentiated_tb_ids", [])),
                "tpt_treatment_start": len(data.get("tpt_treatment_start_ids", [])),
                "tpt_presumptive": len(data.get("tpt_presumptive_ids", [])),
                "adhar_face_auth": len(data.get("adhar_face_authentication_ids", [])),
                "consent_with_id": len(data.get("consent_with_id_ids", [])),
                "culture_dst": len(data.get("culture_dst_ids", [])),
                
                # Raw ID Lists for FO Drill-Down Inspector
                "notification_ids": data.get("notification_ids", []),
                "hiv_dm_ids": data.get("hiv_dm_ids", []),
                "dbt_ids": data.get("dbt_ids", []),
                "sample_collection_ids": data.get("sample_collection_ids", []),
                "sample_tested_ids": data.get("sample_tested_ids", []),
                "outcome_assigned_ids": data.get("outcome_assigned_ids", []),
                "home_visit_ids": data.get("home_visit_ids", []),
                "contact_tracing_ids": data.get("contact_tracing_ids", []),
                "follow_up_ids": data.get("follow_up_ids", []),
                "face_to_face_ids": data.get("face_to_face_ids", []),
                "presumptive_ids": data.get("presumptive_ids", []),
                "documents_ids": data.get("documents_ids", []),
                "fdc_provided_ids": data.get("fdc_provided_ids", []),
                "kit_consumption_ids": data.get("kit_consumption_ids", []),
                "differentiated_tb_ids": data.get("differentiated_tb_ids", []),
                "tpt_treatment_start_ids": data.get("tpt_treatment_start_ids", []),
                "tpt_presumptive_ids": data.get("tpt_presumptive_ids", []),
                "adhar_face_authentication_ids": data.get("adhar_face_authentication_ids", []),
                "consent_with_id_ids": data.get("consent_with_id_ids", []),
                "culture_dst_ids": data.get("culture_dst_ids", []),
                "visited_names": data.get("visited_names", []),
                "remark": data.get("remark", ""),
                
                "is_override": data.get("is_override_used", False)
            })
            
        return {"records": records}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/get-directory")
async def get_directory():
    try:
        docs = db.collection("staff_directory").stream()
        directory = {}
        for doc in docs:
            data = doc.to_dict()
            dist = data.get("district")
            if dist not in directory:
                directory[dist] = []
            directory[dist].append(data.get("name"))
        return directory
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/verify-pin")
async def verify_pin(data: PinCheck):
    try:
        doc_id = f"{data.working_place}_{data.fo_name}".replace(" ", "").lower()
        cache_key = f"pin_{doc_id}"
        cached_pin = cache.get(cache_key)
        
        if cached_pin is not None:
            return {"valid": str(data.pin) == str(cached_pin)}

        staff_doc = await asyncio.to_thread(db.collection("staff_directory").document(doc_id).get)
        
        if not staff_doc.exists:
            return {"valid": False}
            
        real_pin = staff_doc.to_dict().get("pin")
        cache.set(cache_key, str(real_pin), ttl=300) # 5 min cache
        if str(data.pin) == str(real_pin):
            return {"valid": True}
        return {"valid": False}
    except Exception:
        return {"valid": False}

@app.post("/upload-image")
async def upload_image(file: UploadFile = File(...)):
    try:
        bucket = storage.bucket()
        blob = bucket.blob(f"km_photos/{uuid.uuid4()}_{file.filename}")
        blob.upload_from_string(await file.read(), content_type=file.content_type)
        blob.make_public()
        return {"url": blob.public_url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image upload failed: {str(e)}")

class CheckStatusRequest(BaseModel):
    working_place: str
    fo_name: str
    date: str

class StartDayRequest(BaseModel):
    working_place: str
    fo_name: str
    date: str
    morning_km: int
    morning_km_photo_url: str

@app.post("/check-today-status")
async def check_today_status(req: CheckStatusRequest):
    try:
        doc_id = f"{req.working_place}_{req.fo_name}_{req.date}".replace(" ", "_").lower()
        cache_key = f"status_{doc_id}"
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        doc_ref = db.collection("daily_field_reports").document(doc_id)
        doc = await asyncio.to_thread(doc_ref.get)
        
        res = {"status": "not_started"}
        if doc.exists:
            d = doc.to_dict()
            res = {"status": "completed", "submission_count": 1, "data": d}
                
        cache.set(cache_key, res, ttl=20)
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/submit-daily-report")
async def submit_daily_report(report: DailyActivityReport):
    try:
        if not report.date_of_reporting:
            report.date_of_reporting = datetime.now().strftime("%Y-%m-%d")
            
        doc_id = f"{report.working_place}_{report.fo_name}_{report.date_of_reporting}".replace(" ", "_").lower()
        doc_ref = db.collection("daily_field_reports").document(doc_id)
        
        payload = report.dict(exclude_unset=True)
        payload["status"] = "completed"
        payload["timestamp_completed"] = firestore.SERVER_TIMESTAMP
        payload["submission_count"] = 1
        
        doc = doc_ref.get()
        if doc.exists:
            d = doc.to_dict()
            for k, v in payload.items():
                if isinstance(v, list) and k.endswith("_ids"):
                    combined = d.get(k, []) + v
                    payload[k] = list(dict.fromkeys(combined))
                elif k == "visited_names" and isinstance(v, list):
                    combined = d.get(k, []) + v
                    payload[k] = list(dict.fromkeys(combined))
                elif k == "remark" and v:
                    old_remark = d.get("remark", "")
                    if v not in old_remark:
                        payload[k] = f"{old_remark} | {v}".strip(" |")
                    else:
                        payload[k] = old_remark
                        
        doc_ref.set(payload, merge=True)
        cache.delete(f"status_{doc_id}")
        cache.delete_prefix("profile_")
        cache.delete_prefix("dash_")
        cache.delete_prefix("attendance_")
        return {"message": "Daily report submitted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download-excel")
async def download_excel():
    try:
        docs = db.collection("daily_field_reports").stream()
        consolidated_data = []
        
        list_fields_mapping = {
            "notification_ids": "Notification",
            "hiv_dm_ids": "HIV & DM",
            "dbt_ids": "DBT",
            "sample_collection_ids": "Sample Collection",
            "sample_tested_ids": "Sample Tested",
            "outcome_assigned_ids": "Outcome Assigned",
            "home_visit_ids": "Home Visit",
            "contact_tracing_ids": "Contact Tracing",
            "follow_up_ids": "Follow Up",
            "face_to_face_ids": "Face to Face",
            "presumptive_ids": "Presumptive",
            "documents_ids": "Documents",
            "fdc_provided_ids": "FDC Provided",
            "kit_consumption_ids": "Kit Consumption",
            "differentiated_tb_ids": "Differentiated TB",
            "tpt_treatment_start_ids": "TPT Treatment Start",
            "tpt_presumptive_ids": "TPT Presumptive",
            "adhar_face_authentication_ids": "Adhar Face Authentication",
            "consent_with_id_ids": "Consent with ID"
        }
        
        for doc in docs:
            data = doc.to_dict()
            
            # Find the maximum length among all ID arrays
            max_len = 1  # At least 1 row per report
            for key in list_fields_mapping.keys():
                ids = data.get(key) or []
                if len(ids) > max_len:
                    max_len = len(ids)
                    
            for i in range(max_len):
                row = {
                    "Date": data.get("date_of_reporting", ""),
                    "Name": data.get("fo_name", ""),
                    "Designation": data.get("designation", ""),
                    "Block": data.get("working_place", ""),
                }
                
                # Fill array IDs
                for db_key, excel_col in list_fields_mapping.items():
                    ids = data.get(db_key) or []
                    row[excel_col] = ids[i] if i < len(ids) else ""
                    
                # Static data only on first row
                if i == 0:
                    row["Morning KM"] = data.get("morning_km", 0)
                    row["Evening KM"] = data.get("evening_km", 0)
                    row["Total KM"] = data.get("total_km", 0)
                    row["Doctors Visited"] = ", ".join(data.get("visited_names", []))
                    row["Morning KM Photo"] = data.get("morning_km_photo_url", "")
                    row["Evening KM Photo"] = data.get("evening_km_photo_url", "")
                    user_remark = data.get("remark", "")
                    override_str = "[Adjusted]" if data.get("is_override_used") else ""
                    row["Remarks"] = f"{override_str} {user_remark}".strip()
                else:
                    row["Morning KM"] = ""
                    row["Evening KM"] = ""
                    row["Total KM"] = ""
                    row["Doctors Visited"] = ""
                    row["Morning KM Photo"] = ""
                    row["Evening KM Photo"] = ""
                    row["Remarks"] = ""
                    
                consolidated_data.append(row)

        df = pd.DataFrame(consolidated_data)
        df.loc[len(df)] = pd.Series({'Date': 'Designed by Insomniac'})
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Consolidated Report')
        output.seek(0)
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": "attachment; filename=DFY_Consolidated_Report.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



import os
import openpyxl
import re

def safe_filename(district: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", district.strip())
    return cleaned.strip("_") or "UNKNOWN"

def ordinal(n: int) -> str:
    if n == 1:
        return "1ST"
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


class TargetUpdate(BaseModel):
    district: str
    fo_name: str
    target: int
    month: Optional[str] = None

@app.get("/get-targets")
async def get_targets(district: Optional[str] = None, month: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        cache_key = f"targets_{month}_{district or 'all'}_{districts or 'all'}"
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        docs = await asyncio.to_thread(lambda: list(db.collection("staff_targets").stream()))
        
        month_targets = {}
        default_targets = {}
        
        for doc in docs:
            data = doc.to_dict()
            d_dist = data.get("district")
            d_name = data.get("fo_name")
            d_target = int(data.get("target", 50)) if str(data.get("target", "")).isdigit() else 50
            d_month = data.get("month")
            
            if not d_dist or not d_name:
                continue
                
            key = f"{d_dist}_{d_name}".lower()
            
            if d_month == month:
                month_targets[key] = {
                    "fo_name": d_name,
                    "district": d_dist,
                    "target": d_target,
                    "month": month
                }
            elif not d_month:
                default_targets[key] = {
                    "fo_name": d_name,
                    "district": d_dist,
                    "target": d_target,
                    "month": month
                }

        staff_docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        targets = []
        
        for s in staff_docs:
            sd = s.to_dict()
            s_dist = sd.get("district")
            s_name = sd.get("name")
            if not s_dist or not s_name:
                continue
            if allowed_dist_set and s_dist not in allowed_dist_set:
                continue
            if district and district != "All" and s_dist != district:
                continue
                
            key = f"{s_dist}_{s_name}".lower()
            if key in month_targets:
                t_val = month_targets[key]["target"]
            elif key in default_targets:
                t_val = default_targets[key]["target"]
            else:
                t_val = 50
                
            targets.append({
                "fo_name": s_name,
                "district": s_dist,
                "designation": sd.get("designation", "FC"),
                "target": t_val,
                "month": month
            })
            
        targets.sort(key=lambda x: (x["district"], x["fo_name"]))
        res = {"success": True, "month": month, "targets": targets}
        cache.set(cache_key, res, ttl=30)
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/update-target")
async def update_target(data: TargetUpdate):
    try:
        month = data.month or datetime.now().strftime("%Y-%m")
        clean_dist = data.district.strip()
        clean_name = data.fo_name.strip()
        
        # 1. Month-scoped document
        month_doc_id = f"{month}_{clean_dist}_{clean_name}".replace(" ", "").lower()
        await asyncio.to_thread(lambda: db.collection("staff_targets").document(month_doc_id).set({
            "month": month,
            "district": clean_dist,
            "fo_name": clean_name,
            "target": int(data.target),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }, merge=True))
        
        # 2. General fallback document
        fallback_doc_id = f"{clean_dist}_{clean_name}".replace(" ", "").lower()
        await asyncio.to_thread(lambda: db.collection("staff_targets").document(fallback_doc_id).set({
            "district": clean_dist,
            "fo_name": clean_name,
            "target": int(data.target),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }, merge=True))
        
        cache.delete_prefix("targets_")
        cache.delete_prefix("profile_")
        await log_admin_activity(
            action_type="TARGET_UPDATED",
            details=f"Updated target for {clean_name} ({clean_dist}) to {data.target} for month {month}",
            district=clean_dist,
            target_officer=clean_name,
            diff={"month": month, "target": int(data.target)}
        )
        return {"success": True, "month": month, "message": f"Target for {month} successfully updated!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 14 Standard KPI Categories Definition (Exact Master Blueprint)
EXCEL_KPI_CATEGORIES = [
    ("NOTIFICATION", "notification_ids", 3),
    ("HIV & DM", "hiv_dm_ids", 4),
    ("DBT", "dbt_ids", 5),
    ("SAMPLE COLLECTION", "sample_collection_ids", 6),
    ("SAMPLE TESTED", "sample_tested_ids", 7),
    ("Outcome Assigned", "outcome_assigned_ids", 8),
    ("Home Visit", "home_visit_ids", 9),
    ("Contact Tracing", "contact_tracing_ids", 10),
    ("Follow Up", "follow_up_ids", 11),
    ("Face to Face", "face_to_face_ids", 12),
    ("Presumptive", "presumptive_ids", 13),
    ("Documents", "documents_ids", 14),
    ("FDC Provided", "fdc_provided_ids", 15),
    ("Kit Consumption", "kit_consumption_ids", 16)
]

def get_kpi_tab_name(day: int) -> str:
    if day == 1:
        return "1ST"
    elif day == 2:
        return "2nd"
    elif day == 3:
        return "3rd"
    elif day in [21, 31]:
        return f"{day}st"
    elif day == 22:
        return "22nd"
    elif day == 23:
        return "23rd"
    else:
        return f"{day}th"

def generate_district_kpi_bytes(district: str, month_prefix: Optional[str] = None) -> Optional[bytes]:
    if not month_prefix:
        month_prefix = datetime.now().strftime("%Y-%m")
        
    safe_dist = safe_filename(district)
    template_path = f"templates/template_{safe_dist}.xlsx"
    if not os.path.exists(template_path):
        return None
        
    # Load workbook preserving all formulas
    wb = openpyxl.load_workbook(template_path, data_only=False)
    sheet_map = {name.strip().lower(): name for name in wb.sheetnames}
    
    # Configure auto-recalculation
    try:
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    
    # 1. Fetch Targets (Prioritizing Month-Scoped Target)
    target_map = {}
    try:
        t_docs = db.collection("staff_targets").where("district", "==", district).stream()
        for td in t_docs:
            t_data = td.to_dict()
            f_name = re.sub(r'\s+', ' ', str(t_data.get("fo_name", ""))).strip().lower()
            if f_name:
                if t_data.get("month") == month_prefix:
                    target_map[f_name] = int(t_data.get("target", 50))
                elif f_name not in target_map:
                    target_map[f_name] = int(t_data.get("target", 50))
    except Exception as e:
        print(f"Target fetch notice for {district}: {e}")
                    
    # 2. Fetch and Sort Daily Field Reports for this District and Month
    docs = db.collection("daily_field_reports").where("working_place", "==", district).stream()
    reports = []
    for doc in docs:
        d = doc.to_dict()
        date_str = str(d.get("date_of_reporting", "")).strip()
        if date_str and date_str.startswith(month_prefix):
            reports.append(d)
            
    reports.sort(key=lambda x: str(x.get("date_of_reporting", "")))
    
    # Map staff names to their index in this template
    staff_name_to_idx = {}
    ws_day1 = wb["1ST"] if "1ST" in wb.sheetnames else wb[sheet_map.get("1st")] if "1st" in sheet_map else None
    if ws_day1:
        s_idx = 0
        for r in range(2, ws_day1.max_row + 1, 40):
            val = ws_day1.cell(row=r, column=1).value
            if val and str(val).strip():
                staff_name_to_idx[re.sub(r'\s+', ' ', str(val)).strip().lower()] = s_idx
                s_idx += 1

    # Pre-calculate counts for each FO and each KPI
    num_staff = len(staff_name_to_idx)
    num_kpis = len(EXCEL_KPI_CATEGORIES)
    staff_counts = { s_idx: { k_idx: 0 for k_idx in range(num_kpis) } for s_idx in range(num_staff) }
    district_cluster_counts = { c_idx: 0 for c_idx in range(13) }

    # 3. Populate Tabs 3 to 33 ('1ST' to '31st')
    for rep in reports:
        date_str = str(rep.get("date_of_reporting", "")).strip()
        try:
            day_int = int(date_str.split('-')[2])
        except Exception:
            continue
            
        raw_tab_key = get_kpi_tab_name(day_int).lower()
        actual_tab_name = sheet_map.get(raw_tab_key)
        fo_norm = re.sub(r'\s+', ' ', str(rep.get("fo_name", ""))).strip().lower()
        
        if fo_norm in staff_name_to_idx:
            s_idx = staff_name_to_idx[fo_norm]
            
            # Accumulate staff and district counts
            for k_idx, (_, cat_key, _) in enumerate(EXCEL_KPI_CATEGORIES):
                ids = rep.get(cat_key) or []
                if isinstance(ids, list):
                    valid_ids = [str(pid).strip() for pid in ids if str(pid).strip()]
                    staff_counts[s_idx][k_idx] += len(valid_ids)
                    if k_idx < 13:
                        district_cluster_counts[k_idx] += len(valid_ids)

            if actual_tab_name and actual_tab_name in wb.sheetnames:
                ws_day = wb[actual_tab_name]
                start_row = 2 + (s_idx * 40)
                
                # Populate IDs vertically within 40-row bounds
                for kpi_name, cat_key, col_idx in EXCEL_KPI_CATEGORIES:
                    ids = rep.get(cat_key) or []
                    if isinstance(ids, list):
                        valid_ids = [str(pid).strip() for pid in ids if str(pid).strip()]
                        for i in range(min(40, len(valid_ids))):
                            ws_day.cell(row=start_row + i, column=col_idx).value = valid_ids[i]

    # 4. Populate Tab 2: 'CONSOLIDATED SHEET'
    if "CONSOLIDATED SHEET" in wb.sheetnames:
        ws_cons = wb["CONSOLIDATED SHEET"]
        
        # Wing 1: Left Side (District Master Rollup & Master Log) -- Columns A to AM (Cols 1 to 39)
        cluster_row_ptrs = { c_idx: 4 for c_idx in range(13) }
        
        # Write Left Wing Row 2 Grand Totals
        for c_idx in range(13):
            start_c = 1 + (c_idx * 3)
            # Pre-compute exact total count for immediate display across all viewers
            ws_cons.cell(row=2, column=start_c).value = district_cluster_counts[c_idx]

        for rep in reports:
            rep_date = str(rep.get("date_of_reporting", "")).strip()
            rep_fo = str(rep.get("fo_name", "")).strip()
            
            for c_idx in range(13):
                _, cat_key, _ = EXCEL_KPI_CATEGORIES[c_idx]
                ids = rep.get(cat_key) or []
                if isinstance(ids, list):
                    start_c = 1 + (c_idx * 3)
                    for patient_id in ids:
                        pid_str = str(patient_id).strip()
                        if pid_str:
                            r = cluster_row_ptrs[c_idx]
                            c1 = ws_cons.cell(row=r, column=start_c, value=pid_str)
                            c2 = ws_cons.cell(row=r, column=start_c + 1, value=rep_date)
                            c3 = ws_cons.cell(row=r, column=start_c + 2, value=rep_fo)
                            c1.border = EXCEL_THIN_BORDER
                            c2.border = EXCEL_THIN_BORDER
                            c3.border = EXCEL_CLUSTER_DIVIDER
                            c1.alignment = Alignment(horizontal="center", vertical="center")
                            c2.alignment = Alignment(horizontal="center", vertical="center")
                            c3.alignment = Alignment(horizontal="left", vertical="center")
                            cluster_row_ptrs[c_idx] += 1
                            
        # Wing 2: Right Side (Staff-Wise Performance & Indicator Wing) -- Column AN (Col 40) onwards
        staff_kpi_row_ptrs = {}
        for s_idx in range(num_staff):
            staff_base_col = 40 + (s_idx * 14)
            # Write Row 3 Staff Totals
            for k_idx in range(num_kpis):
                staff_kpi_row_ptrs[(s_idx, k_idx)] = 4
                ws_cons.cell(row=3, column=staff_base_col + k_idx).value = staff_counts[s_idx][k_idx]
                
        for rep in reports:
            fo_norm = re.sub(r'\s+', ' ', str(rep.get("fo_name", ""))).strip().lower()
            if fo_norm in staff_name_to_idx:
                s_idx = staff_name_to_idx[fo_norm]
                staff_base_col = 40 + (s_idx * 14)
                
                for k_idx, (_, cat_key, _) in enumerate(EXCEL_KPI_CATEGORIES):
                    ids = rep.get(cat_key) or []
                    if isinstance(ids, list):
                        col = staff_base_col + k_idx
                        for patient_id in ids:
                            pid_str = str(patient_id).strip()
                            if pid_str:
                                r = staff_kpi_row_ptrs[(s_idx, k_idx)]
                                c_cell = ws_cons.cell(row=r, column=col, value=pid_str)
                                if col == staff_base_col + 13:
                                    c_cell.border = EXCEL_CLUSTER_DIVIDER
                                else:
                                    c_cell.border = EXCEL_THIN_BORDER
                                c_cell.alignment = Alignment(horizontal="center", vertical="center")
                                staff_kpi_row_ptrs[(s_idx, k_idx)] += 1

    # 5. Populate Tab 1: 'Performance sheet'
    if "Performance sheet" in wb.sheetnames:
        ws_perf = wb["Performance sheet"]
        for r_idx in range(5, 5 + num_staff):
            cell_name = ws_perf.cell(row=r_idx, column=1).value
            if cell_name and str(cell_name).strip() not in ["GRAND TOTAL", ""]:
                norm_name = re.sub(r'\s+', ' ', str(cell_name)).strip().lower()
                if norm_name in staff_name_to_idx:
                    s_idx = staff_name_to_idx[norm_name]
                    
                    # Col 3: Target
                    target_val = target_map.get(norm_name, 50)
                    ws_perf.cell(row=r_idx, column=3).value = target_val
                    
                    # Col 4: NOTIFICATION
                    notif_count = staff_counts[s_idx][0]
                    ws_perf.cell(row=r_idx, column=4).value = notif_count
                    
                    # Col 5: % Achieved (Formula)
                    cell_pct = ws_perf.cell(row=r_idx, column=5)
                    cell_pct.value = f"=IF(C{r_idx}>0, D{r_idx}/C{r_idx}, 0)"
                    cell_pct.number_format = "0.0%"
                    
                    # Cols 6 to 18: Remaining 13 KPIs
                    for k_idx in range(1, num_kpis):
                        kpi_val = staff_counts[s_idx][k_idx]
                        ws_perf.cell(row=r_idx, column=5 + k_idx).value = kpi_val

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

@app.get("/download-kpi-workbook")
async def download_kpi_workbook(district: str, month: Optional[str] = None):
    try:
        excel_bytes = await asyncio.to_thread(lambda: generate_district_kpi_bytes(district, month))
        if not excel_bytes:
            raise HTTPException(status_code=404, detail=f"Template for {district} not found on server.")
            
        safe_dist = safe_filename(district)
        month_tag = month or datetime.now().strftime("%Y-%m")
        headers = {
            'Content-Disposition': f'attachment; filename="KPI_Report_{safe_dist}_{month_tag}.xlsx"'
        }
        return StreamingResponse(
            io.BytesIO(excel_bytes), 
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download-all-kpi-workbooks")
async def download_all_kpi_workbooks(month: Optional[str] = None, districts: Optional[str] = None):
    try:
        all_bihar = ["Aurangabad", "Begusarai", "Bhojpur", "Buxar", "Darbhanga", "East Champaran", "Gaya", "Jamui", "Jehanabad", "Kaimur", "Khagaria", "Lakhisarai", "Madhubani", "Munger", "Muzaffarpur", "Nawada", "Rohtas", "Samastipur", "Sheikhpura", "Sheohar", "Sitamarhi", "Vaishali"]
        if districts and districts.strip() and districts.strip() != "All":
            allowed_set = set([d.strip() for d in districts.split(",") if d.strip()])
            bihar_districts = [d for d in all_bihar if d in allowed_set]
        else:
            bihar_districts = all_bihar

        zip_buffer = io.BytesIO()
        month_tag = month or datetime.now().strftime("%Y-%m")
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for dist in bihar_districts:
                excel_bytes = await asyncio.to_thread(lambda d=dist: generate_district_kpi_bytes(d, month))
                if excel_bytes:
                    zip_file.writestr(f"KPI_Report_{safe_filename(dist)}_{month_tag}.xlsx", excel_bytes)
                    
        zip_buffer.seek(0)
        archive_name = "DFY_KPI_Scoped_Districts" if (districts and districts != "All") else "DFY_Master_KPI_All_Districts"
        headers = {
            'Content-Disposition': f'attachment; filename="{archive_name}_{month_tag}.zip"'
        }
        return StreamingResponse(
            zip_buffer,
            headers=headers,
            media_type="application/zip"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/staff-directory")
async def get_staff_directory():
    try:
        cached = cache.get("staff_directory_list")
        if cached is not None:
            return {"status": "success", "data": cached}

        docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        directory = {}
        for doc in docs:
            data = doc.to_dict()
            district = data.get("district")
            name = data.get("name")
            if district and name:
                if district not in directory:
                    directory[district] = []
                directory[district].append(name)
        
        for d in directory:
            directory[d] = sorted(directory[d])
            
        cache.set("staff_directory_list", directory, ttl=300) # 5 min cache
        return {"status": "success", "data": directory}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ProfileStatsRequest(BaseModel):
    working_place: str
    fo_name: str
    pin: str
    month: str # format YYYY-MM

@app.post("/my-profile-stats")
async def my_profile_stats(req: ProfileStatsRequest):
    try:
        # Step 1: Verify PIN
        pin_doc = db.collection("staff_directory").document(f"{req.working_place}_{req.fo_name}".replace(" ", "").lower()).get()
        if not pin_doc.exists or pin_doc.to_dict().get("pin") != req.pin:
            raise HTTPException(status_code=401, detail="Invalid PIN")
            
        # Step 2: Fetch Target (Month-Scoped with Fallback)
        target_val = 50
        try:
            req_month = req.month or datetime.now().strftime("%Y-%m")
            m_doc_id = f"{req_month}_{req.working_place}_{req.fo_name}".replace(" ", "").lower()
            m_doc = await asyncio.to_thread(db.collection("staff_targets").document(m_doc_id).get)
            if m_doc.exists:
                target_val = int(m_doc.to_dict().get("target", 50))
            else:
                fb_id = f"{req.working_place}_{req.fo_name}".replace(" ", "").lower()
                fb_doc = await asyncio.to_thread(db.collection("staff_targets").document(fb_id).get)
                if fb_doc.exists:
                    target_val = int(fb_doc.to_dict().get("target", 50))
        except Exception:
            target_val = 50
            
        # Step 3: Fetch all reports for the month
        # Since we don't have indexes for fo_name + date, we can fetch by fo_name and filter in memory
        reports = db.collection("daily_field_reports").where("fo_name", "==", req.fo_name).where("working_place", "==", req.working_place).stream()
        
        stats = {
            "notification": 0,
            "hiv_dm": 0,
            "dbt": 0,
            "sample_collection": 0,
            "sample_tested": 0,
            "outcome_assigned": 0,
            "home_visit": 0,
            "contact_tracing": 0,
            "follow_up": 0,
            "face_to_face": 0,
            "presumptive": 0,
            "fdc_provided": 0,
            "kit_consumption": 0,
            "differentiated_tb": 0,
            "tpt_treatment_start": 0,
            "tpt_presumptive": 0,
            "adhar_face_authentication": 0,
            "consent_with_id": 0,
            "culture_dst": 0
        }
        
        daily_history = {}
        for rep in reports:
            data = rep.to_dict()
            date_str = data.get("date_of_reporting") or data.get("date", "")
            if date_str and date_str.startswith(req.month):
                day_total = 0
                for k in stats.keys():
                    arr = data.get(k + "_ids", [])
                    if isinstance(arr, list):
                        stats[k] += len(arr)
                        day_total += len(arr)
                day_categories = {}
                for k in stats.keys():
                    arr = data.get(k + "_ids", [])
                    if isinstance(arr, list) and len(arr) > 0:
                        day_categories[k] = arr
                        
                daily_history[date_str] = {
                    "submitted": True,
                    "count": data.get("submission_count", 1),
                    "total_ids": day_total,
                    "categories": day_categories,
                    "visited_names": data.get("visited_names", []),
                    "total_km": data.get("total_km", 0),
                    "remark": data.get("remark", "")
                }
                        
        total_achieved = sum(stats.values())
        
        # Calculate Reporting Streak
        sorted_dates = sorted(daily_history.keys(), reverse=True)
        streak_days = 0
        today = datetime.now().date()
        
        # Check streak starting from today or yesterday
        check_date = today
        if today.strftime("%Y-%m-%d") not in daily_history:
            # Maybe today is not yet reported, check from yesterday
            from datetime import timedelta
            check_date = today - timedelta(days=1)
            
        while True:
            d_str = check_date.strftime("%Y-%m-%d")
            if d_str in daily_history and daily_history[d_str].get("submitted"):
                streak_days += 1
                from datetime import timedelta
                check_date = check_date - timedelta(days=1)
            else:
                break

        total_km_month = sum(d.get("total_km", 0) for d in daily_history.values())
        
        badges = []
        if stats.get("notification", 0) >= 100:
            badges.append({"id": "century", "title": "Century Club", "icon": "??", "desc": "100+ Notifications logged"})
        if target_val > 0 and stats.get("notification", 0) >= target_val:
            badges.append({"id": "crusher", "title": "Target Crusher", "icon": "??", "desc": "100% Monthly Target reached"})
        if total_km_month >= 300:
            badges.append({"id": "warrior", "title": "Road Warrior", "icon": "??", "desc": "300+ KM logged this month"})
        if streak_days >= 5:
            badges.append({"id": "streak", "title": "Streak Master", "icon": "??", "desc": f"{streak_days} days continuous reporting"})
        
        return {
            "success": True,
            "target": target_val,
            "total_achieved": total_achieved,
            "breakdown": stats,
            "daily_history": daily_history,
            "streak_days": streak_days,
            "total_km": total_km_month,
            "badges": badges
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/today-attendance")
async def get_today_attendance(date: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")
            
        cache_key = f"attendance_{date}_{districts or 'all'}"
        cached = cache.get(cache_key)
        if cached is not None:
            return cached
            
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        # 1. Fetch all active staff
        staff_docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        staff_list = []
        for doc in staff_docs:
            d = doc.to_dict()
            dist = d.get("district")
            if dist and d.get("name"):
                if allowed_dist_set and dist not in allowed_dist_set:
                    continue
                staff_list.append({
                    "district": dist,
                    "fo_name": d.get("name"),
                    "designation": d.get("designation", "Field Officer")
                })
                
        # 2. Fetch daily field reports for this date
        report_docs = await asyncio.to_thread(lambda: list(db.collection("daily_field_reports").where("date_of_reporting", "==", date).stream()))
        reports_map = {}
        for doc in report_docs:
            d = doc.to_dict()
            dist = d.get('working_place')
            if allowed_dist_set and dist not in allowed_dist_set:
                continue
            key = f"{dist}_{d.get('fo_name')}".replace(" ", "").lower()
            reports_map[key] = {
                "submission_count": d.get("submission_count", 1),
                "total_ids": sum(len(v) for k, v in d.items() if isinstance(v, list) and k.endswith("_ids"))
            }
            
        submitted_full = []
        submitted_partial = []
        missing_fos = []
        
        for s in staff_list:
            key = f"{s['district']}_{s['fo_name']}".replace(" ", "").lower()
            if key in reports_map:
                rep = reports_map[key]
                info = {**s, **rep}
                if rep["submission_count"] >= 2:
                    submitted_full.append(info)
                else:
                    submitted_partial.append(info)
            else:
                missing_fos.append(s)
                
        # Sort missing FOs by district then name
        missing_fos.sort(key=lambda x: (x["district"], x["fo_name"]))
        submitted_full.sort(key=lambda x: (x["district"], x["fo_name"]))
        submitted_partial.sort(key=lambda x: (x["district"], x["fo_name"]))

        res = {
            "date": date,
            "total_staff": len(staff_list),
            "submitted_full_count": len(submitted_full),
            "submitted_partial_count": len(submitted_partial),
            "missing_count": len(missing_fos),
            "submitted_full": submitted_full,
            "submitted_partial": submitted_partial,
            "missing_fos": missing_fos
        }
        cache.set(cache_key, res, ttl=15)
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/duplicate-audit")
async def duplicate_audit(month: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        start_date = f"{month}-01"
        end_date = f"{month}-31"
        
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])
        
        docs = db.collection("daily_field_reports")\
            .where("date_of_reporting", ">=", start_date)\
            .where("date_of_reporting", "<=", end_date)\
            .stream()
            
        id_registry = {} # id -> list of {fo_name, district, date, category}
        
        categories_map = {
            "notification_ids": "Notification",
            "hiv_dm_ids": "HIV & DM",
            "dbt_ids": "DBT",
            "sample_collection_ids": "Sample Collection",
            "sample_tested_ids": "Sample Tested",
            "outcome_assigned_ids": "Outcome Assigned",
            "home_visit_ids": "Home Visit",
            "contact_tracing_ids": "Contact Tracing",
            "follow_up_ids": "Follow Up",
            "face_to_face_ids": "Face to Face",
            "presumptive_ids": "Presumptive",
            "documents_ids": "Documents",
            "fdc_provided_ids": "FDC Provided",
            "kit_consumption_ids": "Kit Consumption",
            "differentiated_tb_ids": "Differentiated TB",
            "tpt_treatment_start_ids": "TPT Treatment Start",
            "tpt_presumptive_ids": "TPT Presumptive",
            "adhar_face_authentication_ids": "Adhar Face Auth",
            "consent_with_id_ids": "Consent with ID",
            "culture_dst_ids": "Culture / DST"
        }
        
        for doc in docs:
            d = doc.to_dict()
            fo = d.get("fo_name", "Unknown")
            dist = d.get("working_place", "Unknown")
            rep_date = d.get("date_of_reporting", "")
            
            for key, label in categories_map.items():
                ids = d.get(key) or []
                if isinstance(ids, list):
                    for patient_id in ids:
                        pid = str(patient_id).strip()
                        if len(pid) >= 5:
                            if pid not in id_registry:
                                id_registry[pid] = []
                            id_registry[pid].append({
                                "fo_name": fo,
                                "district": dist,
                                "date": rep_date,
                                "category": label
                            })
                            
        same_category_duplicates = []
        cross_category_history = []
        
        for pid, occurrences in id_registry.items():
            if len(occurrences) > 1:
                # If district filter is active, at least one occurrence must belong to allowed districts
                if allowed_dist_set and not any(o.get("district") in allowed_dist_set for o in occurrences):
                    continue

                # Check if any category was repeated
                cat_counts = {}
                for o in occurrences:
                    c = o['category']
                    cat_counts[c] = cat_counts.get(c, 0) + 1
                    
                is_same_category = any(cnt > 1 for cnt in cat_counts.values())
                
                entry = {
                    "patient_id": pid,
                    "occurrence_count": len(occurrences),
                    "is_same_category": is_same_category,
                    "repeated_categories": [c for c, cnt in cat_counts.items() if cnt > 1],
                    "occurrences": occurrences
                }
                
                if is_same_category:
                    same_category_duplicates.append(entry)
                else:
                    cross_category_history.append(entry)
                    
        return {
            "status": "success",
            "month": month,
            "total_same_category_duplicates": len(same_category_duplicates),
            "total_cross_category": len(cross_category_history),
            "total_duplicate_ids": len(same_category_duplicates) + len(cross_category_history),
            "same_category_duplicates": same_category_duplicates,
            "cross_category_history": cross_category_history,
            "duplicates": same_category_duplicates + cross_category_history
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Admin Authentication & Zero-Budget Emergency Recovery ---
class AdminLoginReq(BaseModel):
    password: str

class AdminRecoveryReq(BaseModel):
    recovery_code: str
    new_password: str

class AdminChangeSettingsReq(BaseModel):
    current_password: str
    new_password: Optional[str] = None
    new_recovery_key: Optional[str] = None
    new_security_pin: Optional[str] = None

def get_or_init_admin_auth() -> dict:
    doc_ref = db.collection("admin_config").document("auth_settings")
    doc = doc_ref.get()
    if doc.exists:
        return doc.to_dict()
    
    default_auth = {
        "password": "dfyadmin2026",
        "master_recovery_key": "DFY-RESCUE-9921",
        "security_pin": "7788",
        "security_question": "DFY State Organization Code",
        "security_answer": "BIHAR-DFY-TB",
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    doc_ref.set(default_auth)
    return default_auth

@app.post("/admin/auth/login")
async def admin_login(req: AdminLoginReq):
    try:
        auth_data = await asyncio.to_thread(get_or_init_admin_auth)
        correct_pw = auth_data.get("password", "dfyadmin2026")
        if req.password == correct_pw:
            return {"success": True, "message": "Login successful"}
        raise HTTPException(status_code=401, detail="Invalid password")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/auth/settings")
async def get_admin_settings(password: str):
    try:
        auth_data = await asyncio.to_thread(get_or_init_admin_auth)
        if password != auth_data.get("password", "dfyadmin2026"):
            raise HTTPException(status_code=401, detail="Unauthorized")
        return {
            "success": True,
            "master_recovery_key": auth_data.get("master_recovery_key", "DFY-RESCUE-9921"),
            "security_pin": auth_data.get("security_pin", "7788"),
            "last_updated": auth_data.get("last_updated", "")
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/auth/emergency-reset")
async def admin_emergency_reset(req: AdminRecoveryReq):
    try:
        auth_data = await asyncio.to_thread(get_or_init_admin_auth)
        code = req.recovery_code.strip().upper()
        
        valid_key = str(auth_data.get("master_recovery_key", "DFY-RESCUE-9921")).strip().upper()
        valid_pin = str(auth_data.get("security_pin", "7788")).strip()
        valid_ans = str(auth_data.get("security_answer", "BIHAR-DFY-TB")).strip().upper()
        
        if code in [valid_key, valid_pin, valid_ans]:
            doc_ref = db.collection("admin_config").document("auth_settings")
            await asyncio.to_thread(lambda: doc_ref.set({
                "password": req.new_password,
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }, merge=True))
            return {"success": True, "message": "Password successfully reset!"}
        
        raise HTTPException(status_code=400, detail="Invalid Emergency Recovery Key or PIN.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/auth/update-credentials")
async def admin_update_credentials(req: AdminChangeSettingsReq):
    try:
        auth_data = await asyncio.to_thread(get_or_init_admin_auth)
        if req.current_password != auth_data.get("password", "dfyadmin2026"):
            raise HTTPException(status_code=401, detail="Current password incorrect.")
            
        update_payload = {
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        if req.new_password:
            update_payload["password"] = req.new_password
        if req.new_recovery_key:
            update_payload["master_recovery_key"] = req.new_recovery_key
        if req.new_security_pin:
            update_payload["security_pin"] = req.new_security_pin
            
        doc_ref = db.collection("admin_config").document("auth_settings")
        await asyncio.to_thread(lambda: doc_ref.set(update_payload, merge=True))
        return {"success": True, "message": "Admin credentials updated successfully!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/export-state-summary")
async def export_state_summary(month: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        start_date = f"{month}-01"
        end_date = f"{month}-31"
        
        # 1. Fetch reports
        report_docs = await asyncio.to_thread(lambda: list(db.collection("daily_field_reports")
            .where("date_of_reporting", ">=", start_date)
            .where("date_of_reporting", "<=", end_date)
            .stream()))
            
        # 2. Fetch targets
        target_docs = await asyncio.to_thread(lambda: list(db.collection("staff_targets").stream()))
        targets_by_dist = {}
        for td in target_docs:
            d = td.to_dict()
            dist = d.get("district", "Unknown")
            targets_by_dist[dist] = targets_by_dist.get(dist, 0) + (int(d.get("target", 0)) if str(d.get("target", "")).isdigit() else 0)
            
        # 3. Fetch staff count
        staff_docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        staff_by_dist = {}
        for sd in staff_docs:
            dist = sd.to_dict().get("district", "Unknown")
            staff_by_dist[dist] = staff_by_dist.get(dist, 0) + 1
            
        # Aggregate by district
        all_bihar = ["Aurangabad", "Begusarai", "Bhojpur", "Buxar", "Darbhanga", "East Champaran", "Gaya", "Jamui", "Jehanabad", "Kaimur", "Khagaria", "Lakhisarai", "Madhubani", "Munger", "Muzaffarpur", "Nawada", "Rohtas", "Samastipur", "Sheikhpura", "Sheohar", "Sitamarhi", "Vaishali"]
        if districts and districts.strip() and districts.strip() != "All":
            allowed_set = set([d.strip() for d in districts.split(",") if d.strip()])
            bihar_districts = [d for d in all_bihar if d in allowed_set]
        else:
            bihar_districts = all_bihar

        dist_data = {dist: {
            "District": dist,
            "Active Staff": staff_by_dist.get(dist, 0),
            "Monthly Target": targets_by_dist.get(dist, 0),
            "Notifications": 0,
            "Target %": 0,
            "Samples Tested": 0,
            "Presumptive": 0,
            "DBT Seeded": 0,
            "TPT Started": 0,
            "Doctor Visits": 0,
            "Total Travel KM": 0,
            "Reports Submitted": 0
        } for dist in bihar_districts}
        
        for doc in report_docs:
            d = doc.to_dict()
            dist = d.get("working_place", "")
            if dist in dist_data:
                dist_data[dist]["Notifications"] += len(d.get("notification_ids", []))
                dist_data[dist]["Samples Tested"] += len(d.get("sample_tested_ids", []))
                dist_data[dist]["Presumptive"] += len(d.get("presumptive_ids", []))
                dist_data[dist]["DBT Seeded"] += len(d.get("dbt_ids", []))
                dist_data[dist]["TPT Started"] += len(d.get("tpt_treatment_start_ids", []))
                dist_data[dist]["Doctor Visits"] += len(d.get("visited_names", []))
                dist_data[dist]["Total Travel KM"] += int(d.get("total_km", 0) or 0)
                dist_data[dist]["Reports Submitted"] += 1
                
        rows = []
        for dist, data in dist_data.items():
            tgt = data["Monthly Target"]
            ach = data["Notifications"]
            data["Target %"] = f"{round((ach / tgt) * 100)}%" if tgt > 0 else "0%"
            rows.append(data)
            
        df = pd.DataFrame(rows)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="State Performance Summary")
            ws = writer.sheets["State Performance Summary"]
            # Formatting
            style_excel_worksheet(ws, header_fill_color="1E3A8A")
                
        output.seek(0)
        filename = f"DFY_State_Summary_{month}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/export-fo-dossier")
async def export_fo_dossier(month: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        start_date = f"{month}-01"
        end_date = f"{month}-31"
        
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        report_docs = await asyncio.to_thread(lambda: list(db.collection("daily_field_reports")
            .where("date_of_reporting", ">=", start_date)
            .where("date_of_reporting", "<=", end_date)
            .stream()))
            
        staff_docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        staff_map = {}
        for sd in staff_docs:
            d = sd.to_dict()
            dist = d.get("district", "")
            if allowed_dist_set and dist not in allowed_dist_set:
                continue
            key = (dist, d.get("name", ""))
            staff_map[key] = {
                "District": dist,
                "Officer Name": d.get("name", ""),
                "Designation": d.get("designation", "Field Officer"),
                "Active Reporting Days": 0,
                "Total Travel KM": 0,
                "Notifications": 0,
                "Samples Tested": 0,
                "Presumptive": 0,
                "DBT": 0,
                "TPT Start": 0,
                "Doctor Visits": 0,
                "Total All IDs": 0
            }
            
        for doc in report_docs:
            d = doc.to_dict()
            dist = d.get("working_place", "")
            fo = d.get("fo_name", "")
            if allowed_dist_set and dist not in allowed_dist_set:
                continue
            key = (dist, fo)
            if key not in staff_map:
                staff_map[key] = {
                    "District": dist,
                    "Officer Name": fo,
                    "Designation": "Field Officer",
                    "Active Reporting Days": 0,
                    "Total Travel KM": 0,
                    "Notifications": 0,
                    "Samples Tested": 0,
                    "Presumptive": 0,
                    "DBT": 0,
                    "TPT Start": 0,
                    "Doctor Visits": 0,
                    "Total All IDs": 0
                }
                
            entry = staff_map[key]
            entry["Active Reporting Days"] += 1
            entry["Total Travel KM"] += int(d.get("total_km", 0) or 0)
            entry["Notifications"] += len(d.get("notification_ids", []))
            entry["Samples Tested"] += len(d.get("sample_tested_ids", []))
            entry["Presumptive"] += len(d.get("presumptive_ids", []))
            entry["DBT"] += len(d.get("dbt_ids", []))
            entry["TPT Start"] += len(d.get("tpt_treatment_start_ids", []))
            entry["Doctor Visits"] += len(d.get("visited_names", []))
            
            day_total_ids = sum(len(v) for k, v in d.items() if isinstance(v, list) and k.endswith("_ids"))
            entry["Total All IDs"] += day_total_ids
            
        df = pd.DataFrame(list(staff_map.values()))
        if not df.empty:
            df.sort_values(by=["District", "Officer Name"], inplace=True)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="FO Performance Dossier")
            ws = writer.sheets["FO Performance Dossier"]
            # Formatting
            style_excel_worksheet(ws, header_fill_color="047857")
                
        output.seek(0)
        filename = f"DFY_FO_Performance_Dossier_{month}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Patient ID Correction & Editing Suite ---
class EditIdRequest(BaseModel):
    working_place: str
    fo_name: str
    date: str
    category: str # e.g. "notification_ids" or "notification"
    action: str   # "replace", "delete", "add"
    old_id: Optional[str] = ""
    new_id: Optional[str] = ""
    edited_by: Optional[str] = "FO" # "FO" or "Admin"
    pin: Optional[str] = ""

@app.post("/api/reports/edit-id")
async def edit_patient_id(req: EditIdRequest):
    try:
        cat_key = req.category if req.category.endswith("_ids") else f"{req.category}_ids"
        
        if req.action not in ["replace", "delete", "add"]:
            raise HTTPException(status_code=400, detail="Invalid action. Must be 'replace', 'delete', or 'add'.")
            
        if req.action in ["replace", "add"]:
            clean_new_id = str(req.new_id).strip()
            if not clean_new_id.isdigit() or len(clean_new_id) != 9:
                raise HTTPException(status_code=400, detail=f"Invalid Patient ID '{clean_new_id}'. Must be exactly 9 digits.")
            req.new_id = clean_new_id
            
        if req.edited_by == "FO" and req.pin:
            pin_doc_id = f"{req.working_place}_{req.fo_name}".replace(" ", "").lower()
            staff_doc = await asyncio.to_thread(db.collection("staff_directory").document(pin_doc_id).get)
            if staff_doc.exists and str(staff_doc.to_dict().get("pin")) != str(req.pin):
                raise HTTPException(status_code=401, detail="Invalid PIN authorization.")

        doc_id = f"{req.working_place}_{req.fo_name}_{req.date}".replace(" ", "_").lower()
        doc_ref = db.collection("daily_field_reports").document(doc_id)
        doc = await asyncio.to_thread(doc_ref.get)
        
        if not doc.exists:
            docs = await asyncio.to_thread(lambda: list(db.collection("daily_field_reports")
                .where("working_place", "==", req.working_place)
                .where("fo_name", "==", req.fo_name)
                .where("date_of_reporting", "==", req.date)
                .stream()))
            if not docs:
                raise HTTPException(status_code=404, detail="No report found for this date and officer.")
            doc_ref = docs[0].reference
            data = docs[0].to_dict()
        else:
            data = doc.to_dict()

        # ?? Strict 24-Hour Editing Window Rule for Field Officers
        if req.edited_by == "FO":
            is_expired = False
            # Check submitted_at timestamp
            sub_ts = data.get("timestamp") or data.get("submitted_at")
            if sub_ts:
                try:
                    # Parse timestamp format
                    sub_clean = str(sub_ts).replace("Z", "+00:00")
                    if "T" in sub_clean:
                        sub_dt = datetime.fromisoformat(sub_clean).replace(tzinfo=None)
                    else:
                        sub_dt = datetime.strptime(sub_clean, "%Y-%m-%d %H:%M:%S")
                    hours_diff = (datetime.now() - sub_dt).total_seconds() / 3600.0
                    if hours_diff > 24.0:
                        is_expired = True
                except Exception:
                    pass
            
            # Fallback check against date_of_reporting
            if not is_expired:
                try:
                    rep_date = datetime.strptime(req.date, "%Y-%m-%d").date()
                    today = datetime.now().date()
                    if (today - rep_date).days > 1: # More than 1 calendar day ago
                        is_expired = True
                except Exception:
                    pass
                    
            if is_expired:
                raise HTTPException(
                    status_code=403, 
                    detail="Field Officer edit window expired (24 hours limit). 24 ghante beet chuke hain. Kripya badlav ke liye District Admin ya State MIS se sampark karein."
                )
            
        current_list = list(data.get(cat_key, []))
        old_id_clean = str(req.old_id).strip()
        
        if req.action == "replace":
            if old_id_clean not in current_list:
                raise HTTPException(status_code=404, detail=f"Old ID '{old_id_clean}' not found in category '{cat_key}'.")
            idx = current_list.index(old_id_clean)
            current_list[idx] = req.new_id
            
        elif req.action == "delete":
            if old_id_clean not in current_list:
                raise HTTPException(status_code=404, detail=f"ID '{old_id_clean}' not found in category '{cat_key}'.")
            current_list.remove(old_id_clean)
            
        elif req.action == "add":
            if req.new_id in current_list:
                raise HTTPException(status_code=400, detail=f"ID '{req.new_id}' is already present in this category.")
            current_list.append(req.new_id)

        await asyncio.to_thread(lambda: doc_ref.update({
            cat_key: current_list,
            "last_edited_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "last_edited_by": req.edited_by
        }))
        
        log_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "working_place": req.working_place,
            "fo_name": req.fo_name,
            "date": req.date,
            "category": cat_key,
            "action": req.action,
            "old_id": req.old_id,
            "new_id": req.new_id,
            "edited_by": req.edited_by
        }
        await asyncio.to_thread(lambda: db.collection("id_edit_logs").add(log_entry))
        await log_admin_activity(
            action_type=f"PATIENT_ID_{req.action.upper()}",
            details=f"{req.edited_by} {req.action}d ID in {cat_key} for {req.fo_name} on {req.date} (Old: {req.old_id}, New: {req.new_id})",
            district=req.working_place,
            target_officer=req.fo_name,
            user_name=req.edited_by,
            diff={"category": cat_key, "action": req.action, "old_id": req.old_id, "new_id": req.new_id}
        )
        
        cache.delete(f"status_{doc_id}")
        cache.delete_prefix("profile_")
        cache.delete_prefix("dash_")
        cache.delete_prefix("attendance_")
        
        return {
            "success": True,
            "message": f"Patient ID successfully {req.action}d!",
            "category": cat_key,
            "updated_ids": current_list
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Admin Staff & PIN Management Suite ---
class AddStaffReq(BaseModel):
    district: str
    name: str
    pin: str
    designation: Optional[str] = "Field Officer"
    target: Optional[int] = 50

class UpdatePinReq(BaseModel):
    district: str
    name: str
    new_pin: str

class DeleteStaffReq(BaseModel):
    district: str
    name: str

@app.get("/admin/staff/list")
async def get_staff_full_list(districts: Optional[str] = None):
    try:
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        staff = []
        for doc in docs:
            d = doc.to_dict()
            dist = d.get("district")
            if dist and d.get("name"):
                if allowed_dist_set and dist not in allowed_dist_set:
                    continue
                staff.append({
                    "id": doc.id,
                    "district": dist,
                    "name": d.get("name"),
                    "pin": str(d.get("pin", "")),
                    "designation": d.get("designation", "Field Officer"),
                    "created_at": d.get("created_at", "")
                })
        staff.sort(key=lambda s: (s["district"], s["name"]))
        return {"success": True, "staff": staff}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/staff/add")
async def add_staff_member(req: AddStaffReq):
    try:
        clean_dist = req.district.strip()
        clean_name = req.name.strip()
        clean_pin = str(req.pin).strip()
        
        if not clean_dist or not clean_name:
            raise HTTPException(status_code=400, detail="District and Officer Name are required.")
            
        if not clean_pin.isdigit() or len(clean_pin) != 4:
            raise HTTPException(status_code=400, detail="PIN must be exactly 4 digits.")
            
        doc_id = f"{clean_dist}_{clean_name}".replace(" ", "").lower()
        doc_ref = db.collection("staff_directory").document(doc_id)
        
        existing = await asyncio.to_thread(doc_ref.get)
        if existing.exists:
            raise HTTPException(status_code=400, detail=f"Officer '{clean_name}' already exists in '{clean_dist}'.")
            
        payload = {
            "district": clean_dist,
            "name": clean_name,
            "pin": clean_pin,
            "designation": req.designation or "Field Officer",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        await asyncio.to_thread(lambda: doc_ref.set(payload))
        
        target_doc_id = f"{clean_dist}_{clean_name}".replace(" ", "").lower()
        await asyncio.to_thread(lambda: db.collection("staff_targets").document(target_doc_id).set({
            "district": clean_dist,
            "fo_name": clean_name,
            "target": req.target or 50,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }, merge=True))
        
        cache.delete("staff_directory_list")
        cache.delete_prefix("attendance_")
        cache.delete_prefix("targets_")
        
        return {"success": True, "message": f"Officer '{clean_name}' added successfully to {clean_dist}!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/staff/update-pin")
async def update_staff_pin(req: UpdatePinReq):
    try:
        clean_dist = req.district.strip()
        clean_name = req.name.strip()
        clean_pin = str(req.new_pin).strip()
        
        if not clean_pin.isdigit() or len(clean_pin) != 4:
            raise HTTPException(status_code=400, detail="New PIN must be exactly 4 digits.")
            
        doc_id = f"{clean_dist}_{clean_name}".replace(" ", "").lower()
        doc_ref = db.collection("staff_directory").document(doc_id)
        
        doc = await asyncio.to_thread(doc_ref.get)
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Staff record not found.")
            
        await asyncio.to_thread(lambda: doc_ref.update({
            "pin": clean_pin,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }))
        
        cache.delete(f"pin_{doc_id}")
        cache.delete("staff_directory_list")
        
        return {"success": True, "message": f"PIN for '{clean_name}' successfully updated to {clean_pin}!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/staff/delete")
async def delete_staff_member(req: DeleteStaffReq):
    try:
        clean_dist = req.district.strip()
        clean_name = req.name.strip()
        
        doc_id = f"{clean_dist}_{clean_name}".replace(" ", "").lower()
        doc_ref = db.collection("staff_directory").document(doc_id)
        
        doc = await asyncio.to_thread(doc_ref.get)
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Staff record not found.")
            
        await asyncio.to_thread(doc_ref.delete)
        
        cache.delete(f"pin_{doc_id}")
        cache.delete("staff_directory_list")
        cache.delete_prefix("attendance_")
        
        return {"success": True, "message": f"Officer '{clean_name}' removed from directory."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/staff/export-pins")
async def export_staff_pins(district: Optional[str] = "All", districts: Optional[str] = None):
    try:
        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        docs = await asyncio.to_thread(lambda: list(db.collection("staff_directory").stream()))
        rows = []
        s_no = 1
        for doc in docs:
            d = doc.to_dict()
            dist = d.get("district", "")
            name = d.get("name", "")
            if allowed_dist_set and dist not in allowed_dist_set:
                continue
            if district != "All" and dist != district:
                continue
            if dist and name:
                rows.append({
                    "S.No": s_no,
                    "District": dist,
                    "Officer Name": name,
                    "Designation": d.get("designation", "Field Officer"),
                    "4-Digit PIN": str(d.get("pin", "")),
                    "Status": "Active"
                })
                s_no += 1
                
        rows.sort(key=lambda r: (r["District"], r["Officer Name"]))
        for idx, r in enumerate(rows):
            r["S.No"] = idx + 1
            
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_title = f"PINs {district}" if len(district) < 20 else "Staff PINs"
            df.to_excel(writer, index=False, sheet_name=sheet_title[:31])
            ws = writer.sheets[sheet_title[:31]]
            style_excel_worksheet(ws, header_fill_color="1E3A8A")
                
        output.seek(0)
        filename = f"DFY_Staff_PIN_Directory_{district}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Predictive Cascade & Dropout Alerts Engine (Mission Critical Priority) ---
def compute_cascade_alerts(month: str, district: Optional[str] = "All", fo_name: Optional[str] = None, districts: Optional[str] = None):
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    allowed_dist_set = None
    if districts and districts.strip() and districts.strip() != "All":
        allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])
        
    docs = db.collection("daily_field_reports")\
        .where("date_of_reporting", ">=", start_date)\
        .where("date_of_reporting", "<=", end_date)\
        .stream()
        
    patient_map = {}
    
    for doc in docs:
        d = doc.to_dict()
        doc_dist = d.get("working_place", "")
        doc_fo = d.get("fo_name", "")
        doc_date = d.get("date_of_reporting", "")
        
        if allowed_dist_set and doc_dist not in allowed_dist_set:
            continue
        if district != "All" and doc_dist != district:
            continue
        if fo_name and doc_fo != fo_name:
            continue
            
        for cat_key, flag in [
            ("notification_ids", "notification"),
            ("hiv_dm_ids", "hiv_dm"),
            ("dbt_ids", "dbt"),
            ("contact_tracing_ids", "contact_tracing"),
            ("sample_tested_ids", "sample_tested"),
            ("presumptive_ids", "presumptive"),
            ("outcome_assigned_ids", "outcome")
        ]:
            ids = d.get(cat_key, [])
            if isinstance(ids, list):
                for pid in ids:
                    pid_clean = str(pid).strip()
                    if len(pid_clean) >= 5:
                        if pid_clean not in patient_map:
                            patient_map[pid_clean] = {
                                "id": pid_clean,
                                "district": doc_dist,
                                "fo_name": doc_fo,
                                "first_date": doc_date,
                                "notification": False,
                                "hiv_dm": False,
                                "dbt": False,
                                "contact_tracing": False,
                                "sample_tested": False,
                                "presumptive": False,
                                "outcome": False
                            }
                        patient_map[pid_clean][flag] = True
                        if flag in ["notification", "presumptive"] and (not patient_map[pid_clean]["first_date"] or doc_date < patient_map[pid_clean]["first_date"]):
                            patient_map[pid_clean]["first_date"] = doc_date
                            patient_map[pid_clean]["district"] = doc_dist
                            patient_map[pid_clean]["fo_name"] = doc_fo

    alert_list = []
    today_dt = datetime.now().date()
    
    summary = {
        "total_notified": 0,
        "hiv_pending": 0,
        "dbt_pending": 0,
        "contact_pending": 0,
        "udst_pending": 0,
        "presumptive_untested": 0,
        "high_risk_count": 0
    }
    
    for pid, p in patient_map.items():
        days_elapsed = 0
        if p["first_date"]:
            try:
                p_dt = datetime.strptime(p["first_date"], "%Y-%m-%d").date()
                days_elapsed = (today_dt - p_dt).days
            except:
                pass

        # 1. Top Priority: TB Notification Follow-Up Cascade
        if p["notification"]:
            summary["total_notified"] += 1
            missing_actions = []
            
            if not p["hiv_dm"]:
                missing_actions.append("HIV & DM Testing Missing")
                summary["hiv_pending"] += 1
            if not p["dbt"]:
                missing_actions.append("DBT Bank Seeding Missing")
                summary["dbt_pending"] += 1
            if not p["contact_tracing"]:
                missing_actions.append("Contact Tracing Missing")
                summary["contact_pending"] += 1
            if not p["sample_tested"]:
                missing_actions.append("UDST / Testing Missing")
                summary["udst_pending"] += 1
                
            risk_level = "LOW"
            if len(missing_actions) >= 2:
                risk_level = "HIGH"
                summary["high_risk_count"] += 1
            elif len(missing_actions) == 1:
                risk_level = "MEDIUM"
                
            if len(missing_actions) > 0:
                alert_list.append({
                    "id": pid,
                    "district": p["district"],
                    "fo_name": p["fo_name"],
                    "notified_date": p["first_date"],
                    "days_elapsed": days_elapsed,
                    "missing_actions": missing_actions,
                    "risk_level": risk_level,
                    "cascade_type": "Notification",
                    "has_hiv": p["hiv_dm"],
                    "has_dbt": p["dbt"],
                    "has_contact": p["contact_tracing"],
                    "has_udst": p["sample_tested"],
                    "has_outcome": p["outcome"]
                })
                
        # 2. Secondary Priority: Presumptive TB to Testing Cascade
        elif p["presumptive"] and not p["sample_tested"]:
            summary["presumptive_untested"] += 1
            alert_list.append({
                "id": pid,
                "district": p["district"],
                "fo_name": p["fo_name"],
                "notified_date": p["first_date"],
                "days_elapsed": days_elapsed,
                "missing_actions": ["Presumptive TB Testing Pending"],
                "risk_level": "HIGH" if days_elapsed > 7 else "MEDIUM",
                "cascade_type": "Presumptive",
                "has_hiv": p["hiv_dm"],
                "has_dbt": p["dbt"],
                "has_contact": p["contact_tracing"],
                "has_udst": False,
                "has_outcome": p["outcome"]
            })
                
    risk_weight = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
    alert_list.sort(key=lambda x: (risk_weight.get(x["risk_level"], 0), x["days_elapsed"]), reverse=True)
    
    return {"summary": summary, "alerts": alert_list}

@app.get("/api/reports/cascade-alerts")
async def get_cascade_alerts(month: Optional[str] = None, district: Optional[str] = "All", fo_name: Optional[str] = None, districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        cache_key = f"cascade_alerts_{month}_{district}_{fo_name or 'all'}_{districts or 'all'}"
        cached = cache.get(cache_key)
        if cached is not None:
            return cached
            
        data = await asyncio.to_thread(compute_cascade_alerts, month, district, fo_name, districts)
        cache.set(cache_key, data, ttl=180)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/export-cascade-alerts")
async def export_cascade_alerts(month: Optional[str] = None, district: Optional[str] = "All", districts: Optional[str] = None):
    try:
        if not month:
            month = datetime.now().strftime("%Y-%m")
            
        data = await asyncio.to_thread(compute_cascade_alerts, month, district, None, districts)
        alerts = data.get("alerts", [])
        
        rows = []
        for idx, a in enumerate(alerts):
            rows.append({
                "S.No": idx + 1,
                "Patient ID": a["id"],
                "District": a["district"],
                "Field Officer": a["fo_name"],
                "Notification Date": a["notified_date"],
                "Days Elapsed": a["days_elapsed"],
                "Risk Level": a["risk_level"],
                "Missing Interventions": " | ".join(a["missing_actions"]),
                "DBT Status": "Completed" if a["has_dbt"] else "PENDING",
                "HIV/DM Status": "Completed" if a["has_hiv"] else "PENDING",
                "TPT Status": "Completed" if a["has_tpt"] else "PENDING"
            })
            
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_title = f"Cascade Alerts ({district})"
            df.to_excel(writer, index=False, sheet_name=sheet_title[:31])
            ws = writer.sheets[sheet_title[:31]]
            style_excel_worksheet(ws, header_fill_color="B91C1C")
                
        output.seek(0)
        filename = f"DFY_Cascade_Dropout_Alerts_{district}_{month}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =========================================================================
# --- Enterprise Multi-Admin RBAC & Activity Audit Trail Suite ---
# =========================================================================

class AdminUserLoginReq(BaseModel):
    username: str
    password: str

class AdminUserCreateReq(BaseModel):
    username: str
    name: str
    password: str
    role: Optional[str] = "SUB_ADMIN" # "SUPER_ADMIN" or "SUB_ADMIN"
    allowed_districts: Optional[List[str]] = ["All"]
    permissions: Optional[Dict[str, bool]] = {
        "can_view_dashboard": True,
        "can_edit_targets": False,
        "can_manage_staff": False,
        "can_edit_patient_ids": False,
        "can_export_reports": True,
        "can_view_audit_logs": False
    }
    status: Optional[str] = "ACTIVE"
    created_by: Optional[str] = "Super Admin"

class AdminUserUpdateReq(BaseModel):
    user_id: str
    name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    allowed_districts: Optional[List[str]] = None
    permissions: Optional[Dict[str, bool]] = None
    status: Optional[str] = None

class AuditLogQueryReq(BaseModel):
    action_type: Optional[str] = "All"
    district: Optional[str] = "All"
    user_id: Optional[str] = "All"
    search: Optional[str] = ""
    limit: Optional[int] = 200

async def log_admin_activity(
    action_type: str,
    details: str,
    user_name: str = "Super Admin",
    user_id: str = "admin",
    role: str = "SUPER_ADMIN",
    district: Optional[str] = "All",
    target_officer: Optional[str] = "",
    diff: Optional[Dict[str, Any]] = None,
    ip_address: Optional[str] = ""
):
    try:
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action_type": action_type,
            "details": details,
            "user_name": user_name,
            "user_id": user_id,
            "role": role,
            "district": district or "All",
            "target_officer": target_officer or "",
            "diff": diff or {},
            "ip_address": ip_address or ""
        }
        await asyncio.to_thread(lambda: db.collection("admin_audit_logs").add(entry))
    except Exception as e:
        print(f"Audit log background notice: {e}")

async def init_default_super_admin():
    try:
        doc_ref = db.collection("admin_users").document("admin")
        doc = await asyncio.to_thread(doc_ref.get)
        if not doc.exists:
            default_super = {
                "user_id": "admin",
                "username": "admin",
                "name": "Super Admin (Master)",
                "password": "dfyadmin2026",
                "role": "SUPER_ADMIN",
                "allowed_districts": ["All"],
                "permissions": {
                    "can_view_dashboard": True,
                    "can_edit_targets": True,
                    "can_manage_staff": True,
                    "can_edit_patient_ids": True,
                    "can_export_reports": True,
                    "can_view_audit_logs": True
                },
                "status": "ACTIVE",
                "created_by": "System Root",
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "last_login": ""
            }
            await asyncio.to_thread(lambda: doc_ref.set(default_super))
    except Exception as e:
        print(f"Super admin init notice: {e}")

@app.post("/admin/auth/user-login")
async def admin_user_login(req: AdminUserLoginReq):
    try:
        await init_default_super_admin()
        clean_user = req.username.strip().lower()
        
        # Check in admin_users collection
        user_doc_ref = db.collection("admin_users").document(clean_user)
        user_doc = await asyncio.to_thread(user_doc_ref.get)
        
        if not user_doc.exists:
            # Fallback check for query by username
            docs = await asyncio.to_thread(lambda: list(db.collection("admin_users").where("username", "==", clean_user).stream()))
            if docs:
                user_doc = docs[0]
            else:
                # Master legacy password fallback
                auth_data = await asyncio.to_thread(get_or_init_admin_auth)
                if req.password == auth_data.get("password", "dfyadmin2026") and clean_user in ["admin", "superadmin", "dfyadmin"]:
                    user_data = {
                        "user_id": "admin",
                        "username": "admin",
                        "name": "Super Admin",
                        "role": "SUPER_ADMIN",
                        "allowed_districts": ["All"],
                        "permissions": {
                            "can_view_dashboard": True,
                            "can_edit_targets": True,
                            "can_manage_staff": True,
                            "can_edit_patient_ids": True,
                            "can_export_reports": True,
                            "can_view_audit_logs": True
                        },
                        "status": "ACTIVE"
                    }
                    await log_admin_activity("LOGIN_SUCCESS", "Super Admin master login", user_name="Super Admin", user_id="admin", role="SUPER_ADMIN")
                    return {"success": True, "user": user_data}
                
                await log_admin_activity("LOGIN_FAILED", f"Failed login attempt for username '{req.username}'", user_name=req.username, user_id=clean_user, role="UNKNOWN")
                raise HTTPException(status_code=401, detail="Invalid username or password.")
                
        user_data = user_doc.to_dict()
        if user_data.get("status") != "ACTIVE":
            raise HTTPException(status_code=403, detail="Your admin account has been disabled. Contact Super Admin.")
            
        if user_data.get("password") != req.password:
            await log_admin_activity("LOGIN_FAILED", f"Incorrect password for user '{clean_user}'", user_name=user_data.get("name", clean_user), user_id=clean_user, role=user_data.get("role", "SUB_ADMIN"))
            raise HTTPException(status_code=401, detail="Invalid username or password.")
            
        # Update last login timestamp
        await asyncio.to_thread(lambda: user_doc.reference.update({"last_login": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}))
        
        # Don't return password in payload
        safe_user = {k: v for k, v in user_data.items() if k != "password"}
        await log_admin_activity("LOGIN_SUCCESS", f"User {user_data.get('name')} logged in successfully", user_name=user_data.get("name"), user_id=clean_user, role=user_data.get("role", "SUB_ADMIN"))
        
        return {"success": True, "user": safe_user}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/users/list")
async def list_admin_users():
    try:
        await init_default_super_admin()
        docs = await asyncio.to_thread(lambda: list(db.collection("admin_users").stream()))
        users = []
        for doc in docs:
            d = doc.to_dict()
            safe_d = {k: v for k, v in d.items() if k != "password"}
            users.append(safe_d)
            
        users.sort(key=lambda x: (x.get("role") != "SUPER_ADMIN", x.get("name", "")))
        return {"success": True, "users": users}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/users/create")
async def create_admin_user(req: AdminUserCreateReq):
    try:
        clean_user = req.username.strip().lower()
        if not clean_user or not req.password:
            raise HTTPException(status_code=400, detail="Username and password are required.")
            
        doc_ref = db.collection("admin_users").document(clean_user)
        existing = await asyncio.to_thread(doc_ref.get)
        if existing.exists:
            raise HTTPException(status_code=400, detail=f"Username '{clean_user}' is already taken.")
            
        new_user = {
            "user_id": clean_user,
            "username": clean_user,
            "name": req.name.strip(),
            "password": req.password,
            "role": req.role or "SUB_ADMIN",
            "allowed_districts": req.allowed_districts or ["All"],
            "permissions": req.permissions or {
                "can_view_dashboard": True,
                "can_edit_targets": False,
                "can_manage_staff": False,
                "can_edit_patient_ids": False,
                "can_export_reports": True,
                "can_view_audit_logs": False
            },
            "status": req.status or "ACTIVE",
            "created_by": req.created_by or "Super Admin",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "last_login": ""
        }
        await asyncio.to_thread(lambda: doc_ref.set(new_user))
        await log_admin_activity("ADMIN_USER_CREATED", f"Created new admin account '{clean_user}' ({req.name}) with role {req.role}", user_name=req.created_by, role="SUPER_ADMIN")
        
        safe_user = {k: v for k, v in new_user.items() if k != "password"}
        return {"success": True, "user": safe_user, "message": f"User {req.name} successfully created!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/users/update")
async def update_admin_user(req: AdminUserUpdateReq):
    try:
        clean_user = req.user_id.strip().lower()
        doc_ref = db.collection("admin_users").document(clean_user)
        doc = await asyncio.to_thread(doc_ref.get)
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"Admin user '{clean_user}' not found.")
            
        update_data = {"updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        if req.name is not None:
            update_data["name"] = req.name.strip()
        if req.password:
            update_data["password"] = req.password
        if req.role is not None:
            update_data["role"] = req.role
        if req.allowed_districts is not None:
            update_data["allowed_districts"] = req.allowed_districts
        if req.permissions is not None:
            update_data["permissions"] = req.permissions
        if req.status is not None:
            update_data["status"] = req.status
            
        await asyncio.to_thread(lambda: doc_ref.update(update_data))
        await log_admin_activity("PERMISSIONS_UPDATED", f"Updated settings/permissions for admin user '{clean_user}'", user_name="Super Admin", role="SUPER_ADMIN")
        return {"success": True, "message": f"User {clean_user} updated successfully!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/users/delete")
async def delete_admin_user(user_id: str):
    try:
        clean_user = user_id.strip().lower()
        if clean_user == "admin":
            raise HTTPException(status_code=400, detail="Cannot delete master root admin account.")
            
        doc_ref = db.collection("admin_users").document(clean_user)
        await asyncio.to_thread(doc_ref.delete)
        await log_admin_activity("ADMIN_USER_DELETED", f"Deleted admin user account '{clean_user}'", user_name="Super Admin", role="SUPER_ADMIN")
        return {"success": True, "message": f"User {clean_user} deleted successfully!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/audit-logs")
async def get_audit_logs(query: AuditLogQueryReq):
    try:
        # Fetch audit logs ordered chronologically descending
        docs = await asyncio.to_thread(lambda: list(db.collection("admin_audit_logs")
            .order_by("timestamp", direction=firestore.Query.DESCENDING)
            .limit(query.limit or 200)
            .stream()))
            
        logs = []
        for doc in docs:
            d = doc.to_dict()
            
            # Apply filters in memory
            if query.action_type and query.action_type != "All" and d.get("action_type") != query.action_type:
                continue
            if query.district and query.district != "All" and d.get("district") != query.district:
                continue
            if query.user_id and query.user_id != "All" and d.get("user_id") != query.user_id:
                continue
            if query.search:
                s_lower = query.search.lower()
                text_to_search = f"{d.get('details', '')} {d.get('user_name', '')} {d.get('target_officer', '')} {d.get('district', '')}".lower()
                if s_lower not in text_to_search:
                    continue
                    
            logs.append(d)
            
        return {"success": True, "total": len(logs), "logs": logs}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/export-audit-logs")
async def export_audit_logs(action_type: Optional[str] = "All", district: Optional[str] = "All"):
    try:
        docs = await asyncio.to_thread(lambda: list(db.collection("admin_audit_logs")
            .order_by("timestamp", direction=firestore.Query.DESCENDING)
            .limit(1000)
            .stream()))
            
        rows = []
        for idx, doc in enumerate(docs):
            d = doc.to_dict()
            if action_type and action_type != "All" and d.get("action_type") != action_type:
                continue
            if district and district != "All" and d.get("district") != district:
                continue
                
            rows.append({
                "S.No": idx + 1,
                "Timestamp": d.get("timestamp", ""),
                "Admin User": d.get("user_name", ""),
                "Role": d.get("role", ""),
                "Action Type": d.get("action_type", ""),
                "District": d.get("district", ""),
                "Target Officer": d.get("target_officer", ""),
                "Activity Details": d.get("details", ""),
                "IP Address": d.get("ip_address", "")
            })
            
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Admin Audit Trail")
            ws = writer.sheets["Admin Audit Trail"]
            style_excel_worksheet(ws, header_fill_color="1E293B")
            
        output.seek(0)
        filename = f"DFY_Admin_Audit_Trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Enterprise Broadcast & Urgent Announcements Engine ---
class BroadcastCreateReq(BaseModel):
    title: str
    message: str
    priority: Optional[str] = "MEDIUM" # HIGH, MEDIUM, INFO
    target_audience: Optional[str] = "ALL" # ALL, FIELD_STAFF, SUB_ADMINS
    target_districts: Optional[List[str]] = ["All"] # ["All"] or ["Buxar", "Bhojpur", ...]
    created_by_user: Optional[str] = "Super Admin"
    created_by_role: Optional[str] = "SUPER_ADMIN"
    allowed_districts: Optional[List[str]] = None

class BroadcastDeleteReq(BaseModel):
    broadcast_id: str
    requested_by_user: Optional[str] = "admin"
    requested_by_role: Optional[str] = "SUPER_ADMIN"
    allowed_districts: Optional[List[str]] = None

@app.post("/api/broadcasts/create")
async def create_broadcast(req: BroadcastCreateReq):
    try:
        clean_title = req.title.strip()
        clean_msg = req.message.strip()
        if not clean_title or not clean_msg:
            raise HTTPException(status_code=400, detail="Title and message content are required.")

        role = (req.created_by_role or "SUPER_ADMIN").upper()
        # RBAC Check for Sub-Admin:
        target_dists = req.target_districts or ["All"]
        if role == "SUB_ADMIN":
            user_allowed = req.allowed_districts or []
            if "All" in target_dists:
                # Sub-admin cannot broadcast to 'All' Bihar districts, must be scoped to user_allowed
                target_dists = [d for d in user_allowed if d != "All"]
            else:
                # Ensure all selected districts are in user_allowed
                target_dists = [d for d in target_dists if d in user_allowed]
            if not target_dists:
                raise HTTPException(status_code=403, detail="Sub-Admins can only broadcast to their assigned districts.")

        broadcast_id = f"bc_{int(datetime.now().timestamp() * 1000)}"
        doc_data = {
            "id": broadcast_id,
            "title": clean_title,
            "message": clean_msg,
            "priority": (req.priority or "MEDIUM").upper(),
            "target_audience": (req.target_audience or "ALL").upper(),
            "target_districts": target_dists,
            "created_by_user": req.created_by_user or "Admin",
            "created_by_role": role,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "is_active": True
        }

        await asyncio.to_thread(lambda: db.collection("broadcast_alerts").document(broadcast_id).set(doc_data))
        cache.delete_prefix("broadcasts_")

        await log_admin_activity(
            action_type="BROADCAST_CREATED",
            details=f"Created [{req.priority}] broadcast: '{clean_title}' for {', '.join(target_dists)} ({req.target_audience})",
            district=target_dists[0] if len(target_dists) == 1 else "Statewide",
            user_name=req.created_by_user,
            role=role
        )

        return {"success": True, "message": "Broadcast created successfully!", "broadcast": doc_data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/broadcasts/active")
async def get_active_broadcasts(
    district: Optional[str] = None, 
    role: Optional[str] = None, # 'FIELD_STAFF' or 'SUB_ADMIN'
    districts: Optional[str] = None
):
    try:
        cache_key = f"broadcasts_active_{district or 'all'}_{role or 'all'}_{districts or 'all'}"
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        docs = await asyncio.to_thread(lambda: list(db.collection("broadcast_alerts")
            .where("is_active", "==", True)
            .stream()))

        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        active_list = []
        for doc in docs:
            d = doc.to_dict()
            target_aud = d.get("target_audience", "ALL").upper()
            target_dists = d.get("target_districts", ["All"])

            # 1. Audience Filter
            if role:
                r_upper = role.upper()
                if r_upper == "FIELD_STAFF" and target_aud not in ["ALL", "FIELD_STAFF"]:
                    continue
                if r_upper == "SUB_ADMIN" and target_aud not in ["ALL", "SUB_ADMINS"]:
                    continue

            # 2. District Filter
            # If specific district is passed (e.g. Field Officer in Buxar):
            if district and district != "All":
                if "All" not in target_dists and district not in target_dists:
                    continue

            # If multi-district list is passed (e.g. Sub-Admin with Buxar, Bhojpur):
            if allowed_dist_set:
                if "All" not in target_dists and not any(td in allowed_dist_set for td in target_dists):
                    continue

            active_list.append(d)

        active_list.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        res = {"success": True, "broadcasts": active_list}
        cache.set(cache_key, res, ttl=10) # 10 seconds cache for fast broadcast updates
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/broadcasts/all")
async def get_all_broadcasts(districts: Optional[str] = None, role: Optional[str] = None):
    try:
        docs = await asyncio.to_thread(lambda: list(db.collection("broadcast_alerts")
            .order_by("created_at", direction=firestore.Query.DESCENDING)
            .limit(100)
            .stream()))

        allowed_dist_set = None
        if districts and districts.strip() and districts.strip() != "All":
            allowed_dist_set = set([d.strip() for d in districts.split(",") if d.strip()])

        broadcasts = []
        for doc in docs:
            d = doc.to_dict()
            target_dists = d.get("target_districts", ["All"])
            if allowed_dist_set:
                if "All" not in target_dists and not any(td in allowed_dist_set for td in target_dists):
                    continue
            broadcasts.append(d)

        return {"success": True, "broadcasts": broadcasts}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/broadcasts/delete")
async def delete_broadcast(req: BroadcastDeleteReq):
    try:
        doc_ref = db.collection("broadcast_alerts").document(req.broadcast_id)
        doc = await asyncio.to_thread(doc_ref.get)
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Broadcast not found.")

        d = doc.to_dict()
        user_role = (req.requested_by_role or "SUPER_ADMIN").upper()

        if user_role != "SUPER_ADMIN":
            # Sub-admin can only delete if they created it or it matches their allowed districts
            created_by = d.get("created_by_user", "")
            if created_by != req.requested_by_user:
                target_dists = d.get("target_districts", [])
                allowed = req.allowed_districts or []
                if not any(td in allowed for td in target_dists):
                    raise HTTPException(status_code=403, detail="Permission denied to delete this broadcast.")

        await asyncio.to_thread(doc_ref.delete)
        cache.delete_prefix("broadcasts_")

        await log_admin_activity(
            action_type="BROADCAST_DELETED",
            details=f"Deleted broadcast '{d.get('title')}': {req.broadcast_id}",
            district="Statewide" if "All" in d.get("target_districts", []) else d.get("target_districts", [""])[0],
            user_name=req.requested_by_user,
            role=user_role
        )

        return {"success": True, "message": "Broadcast deleted successfully!"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
