import openpyxl
import glob

for f in sorted(glob.glob("templates/*.xlsx")):
    wb = openpyxl.load_workbook(f, data_only=True)
    # find 1st or 1ST sheet
    target_sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("1"):
            target_sheet = wb[name]
            break
    if target_sheet:
        fo_rows = []
        for r in range(1, target_sheet.max_row + 1):
            val = target_sheet.cell(row=r, column=1).value
            if val and val != "NAME":
                fo_rows.append((r, val))
        print(f"{f}: Found {len(fo_rows)} FOs in Day 1 sheet -> {fo_rows[:3]}...")
