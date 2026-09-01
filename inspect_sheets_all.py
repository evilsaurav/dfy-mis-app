import openpyxl

wb = openpyxl.load_workbook("templates/template_Jamui.xlsx")

if "Performance sheet" in wb.sheetnames:
    ws_p = wb["Performance sheet"]
    print("--- Performance sheet (First 10 rows) ---")
    for r in range(1, 10):
        row_vals = [ws_p.cell(row=r, column=c).value for c in range(1, 15)]
        print(f"Row {r:2d}: {row_vals}")

if "CONSOLIDATED SHEET" in wb.sheetnames:
    ws_c = wb["CONSOLIDATED SHEET"]
    print("\n--- CONSOLIDATED SHEET Headers ---")
    headers = [ws_c.cell(row=1, column=c).value for c in range(1, ws_c.max_column + 1)]
    print("Col 1 to 25:", headers[:25])
    print("Col 26 to end:", headers[25:])
