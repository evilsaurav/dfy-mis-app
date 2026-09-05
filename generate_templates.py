# -*- coding: utf-8 -*-
"""
generate_templates.py
Master Blueprint: District KPI Multi-Tab Excel Engine Template Generator
Creates 33-tab production-grade KPI templates for all 10 Bihar districts with professional borders and formatting.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import csv

def load_district_staff():
    mapping = {}
    csv_path = os.path.join(os.path.dirname(__file__), "staff_master.csv")
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r in reader:
                d = r.get("District", "").strip()
                n = r.get("Name", "").strip()
                if d and n:
                    if d not in mapping:
                        mapping[d] = []
                    if n not in mapping[d]:
                        mapping[d].append(n)
    return mapping

DISTRICT_STAFF = load_district_staff()

KPI_CATEGORIES = [
    ("NOTIFICATION", "notification_ids"),
    ("HIV & DM", "hiv_dm_ids"),
    ("DBT", "dbt_ids"),
    ("SAMPLE COLLECTION", "sample_collection_ids"),
    ("SAMPLE TESTED", "sample_tested_ids"),
    ("Outcome Assigned", "outcome_assigned_ids"),
    ("Home Visit", "home_visit_ids"),
    ("Contact Tracing", "contact_tracing_ids"),
    ("Follow Up", "follow_up_ids"),
    ("Face to Face", "face_to_face_ids"),
    ("Presumptive", "presumptive_ids"),
    ("Documents", "documents_ids"),
    ("FDC Provided", "fdc_provided_ids"),
    ("Kit Consumption", "kit_consumption_ids")
]

def get_ordinal_tab_name(day: int) -> str:
    if day == 1:
        return "1ST"
    elif day == 2:
        return "2nd"
    elif day == 3:
        return "3rd"
    elif day in [21, 31]:
        return f"{day}st"
    elif day == 22:
        return "22nd"
    elif day == 23:
        return "23rd"
    else:
        return f"{day}th"

# Styling definitions
font_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_regular = Font(name="Calibri", size=10, bold=False, color="000000")

fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
fill_indigo = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
fill_gray_header = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
fill_green_header = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
fill_gold_total = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

# Border definitions
thin_cell_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

header_border = Border(
    left=Side(style='thin', color='94A3B8'),
    right=Side(style='thin', color='94A3B8'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='medium', color='1E293B')
)

cluster_divider_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='medium', color='64748B'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

cluster_header_divider_border = Border(
    left=Side(style='thin', color='94A3B8'),
    right=Side(style='medium', color='FFFFFF'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='medium', color='1E293B')
)

block_bottom_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='medium', color='475569')
)

total_row_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='double', color='1E293B')
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def generate_district_template(district: str, staff_list: list, output_path: str):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    daily_tab_names = [get_ordinal_tab_name(d) for d in range(1, 32)]
    
    # -------------------------------------------------------------
    # 1. TAB 1: 'Performance sheet'
    # -------------------------------------------------------------
    ws_perf = wb.create_sheet(title="Performance sheet")
    ws_perf.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_perf.merge_cells("A1:R2")
    ws_perf["A1"] = f"DOCTORS FOR YOU (DFY) -- DISTRICT KPI PERFORMANCE SHEET ({district.upper()})"
    ws_perf["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_perf["A1"].fill = fill_navy
    ws_perf["A1"].alignment = align_center
    for r in range(1, 3):
        for c in range(1, 19):
            ws_perf.cell(row=r, column=c).border = header_border
    
    # Subtitle Row 3
    ws_perf["A3"] = f"Monthly Evaluation & Notification Target Achievement Matrix"
    ws_perf["A3"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    
    # Row 4: Headers
    perf_headers = [
        "Employee Name", "DESIG.", "Target", "NOTIFICATION", "% Achieved"
    ] + [kpi[0] for kpi in KPI_CATEGORIES[1:]] # HIV & DM to Kit Consumption
    
    for c_idx, h in enumerate(perf_headers, start=1):
        cell = ws_perf.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_indigo if c_idx <= 5 else fill_gray_header
        cell.alignment = align_center
        cell.border = header_border
        
    # Row 5 onwards: Staff Rows
    start_r = 5
    for idx, staff_name in enumerate(staff_list):
        r = start_r + idx
        cell_name = ws_perf.cell(row=r, column=1, value=staff_name)
        cell_name.font = font_bold
        cell_name.alignment = align_left
        
        cell_desig = ws_perf.cell(row=r, column=2, value="Field Officer")
        cell_desig.font = font_regular
        cell_desig.alignment = align_center
        
        cell_tgt = ws_perf.cell(row=r, column=3, value=50) # Target
        cell_tgt.font = font_bold
        cell_tgt.alignment = align_center
        
        cons_start_col = 40 + (idx * 14)
        notif_col_letter = get_column_letter(cons_start_col)
        
        cell_notif = ws_perf.cell(row=r, column=4, value=0)
        cell_notif.font = font_bold
        cell_notif.alignment = align_center
        
        cell_pct = ws_perf.cell(row=r, column=5, value=f"=IF(C{r}>0, D{r}/C{r}, 0)")
        cell_pct.font = font_bold
        cell_pct.alignment = align_center
        cell_pct.number_format = "0.0%"
        
        for k_idx in range(1, len(KPI_CATEGORIES)):
            col_target = 5 + k_idx
            c = ws_perf.cell(row=r, column=col_target, value=0)
            c.font = font_regular
            c.alignment = align_center
            c.border = thin_cell_border
            
        for c in range(1, len(perf_headers) + 1):
            ws_perf.cell(row=r, column=c).border = thin_cell_border
            
    # Grand Total Row
    total_r = start_r + len(staff_list)
    cell_gt = ws_perf.cell(row=total_r, column=1, value="GRAND TOTAL")
    cell_gt.font = font_title
    cell_gt.fill = fill_navy
    cell_gt.alignment = align_left
    
    ws_perf.cell(row=total_r, column=2, value="").fill = fill_navy
    
    # Target Total
    c_tgt = ws_perf.cell(row=total_r, column=3, value=f"=SUM(C{start_r}:C{total_r-1})")
    c_tgt.font = font_title
    c_tgt.fill = fill_navy
    c_tgt.alignment = align_center
    
    # Notif Total
    c_notif = ws_perf.cell(row=total_r, column=4, value=f"=SUM(D{start_r}:D{total_r-1})")
    c_notif.font = font_title
    c_notif.fill = fill_navy
    c_notif.alignment = align_center
    
    # Overall % Achieved Total
    cell_tot_pct = ws_perf.cell(row=total_r, column=5, value=f"=IF(C{total_r}>0, D{total_r}/C{total_r}, 0)")
    cell_tot_pct.font = font_title
    cell_tot_pct.fill = fill_navy
    cell_tot_pct.alignment = align_center
    cell_tot_pct.number_format = "0.0%"
    
    for k_idx in range(1, len(KPI_CATEGORIES)):
        col_target = 5 + k_idx
        col_letter = get_column_letter(col_target)
        c = ws_perf.cell(row=total_r, column=col_target, value=f"=SUM({col_letter}{start_r}:{col_letter}{total_r-1})")
        c.font = font_title
        c.fill = fill_navy
        c.alignment = align_center
        c.border = total_row_border
        
    for c in range(1, len(perf_headers) + 1):
        ws_perf.cell(row=total_r, column=c).border = total_row_border

    ws_perf.column_dimensions['A'].width = 24
    ws_perf.column_dimensions['B'].width = 14
    ws_perf.column_dimensions['C'].width = 12
    ws_perf.column_dimensions['D'].width = 15
    ws_perf.column_dimensions['E'].width = 14
    for c in range(6, len(perf_headers) + 1):
        ws_perf.column_dimensions[get_column_letter(c)].width = 16

    # -------------------------------------------------------------
    # 2. TAB 2: 'CONSOLIDATED SHEET'
    # -------------------------------------------------------------
    ws_cons = wb.create_sheet(title="CONSOLIDATED SHEET")
    ws_cons.views.sheetView[0].showGridLines = True
    
    # Wing 1: Left Side (District Master Rollup & Master Log) -- Columns A to AM (Cols 1 to 39)
    left_clusters = [kpi[0] for kpi in KPI_CATEGORIES[:13]]
    
    for c_idx, cluster_name in enumerate(left_clusters):
        start_c = 1 + (c_idx * 3)
        end_c = start_c + 2
        
        ws_cons.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
        r1_cell = ws_cons.cell(row=1, column=start_c, value=cluster_name)
        r1_cell.font = font_title
        r1_cell.fill = fill_navy
        r1_cell.alignment = align_center
        
        c_letter = get_column_letter(start_c)
        ws_cons.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)
        r2_cell = ws_cons.cell(row=2, column=start_c, value=0)
        r2_cell.font = font_title
        r2_cell.fill = fill_indigo
        r2_cell.alignment = align_center
        
        subheaders = ["Patient ID", "Date", "Reported by"]
        for s_idx, sh in enumerate(subheaders):
            sub_cell = ws_cons.cell(row=3, column=start_c + s_idx, value=sh)
            sub_cell.font = font_header
            sub_cell.fill = fill_gray_header
            sub_cell.alignment = align_center
            
        for r in range(1, 4):
            for c in range(start_c, end_c + 1):
                if c == end_c:
                    ws_cons.cell(row=r, column=c).border = cluster_divider_border
                else:
                    ws_cons.cell(row=r, column=c).border = thin_cell_border
                    
        # Pre-border initial 50 rows of data
        for r in range(4, 54):
            for c in range(start_c, end_c + 1):
                if c == end_c:
                    ws_cons.cell(row=r, column=c).border = cluster_divider_border
                else:
                    ws_cons.cell(row=r, column=c).border = thin_cell_border
                    
        ws_cons.column_dimensions[get_column_letter(start_c)].width = 14
        ws_cons.column_dimensions[get_column_letter(start_c+1)].width = 12
        ws_cons.column_dimensions[get_column_letter(start_c+2)].width = 18

    # Wing 2: Right Side (Staff-Wise Performance & Indicator Wing) -- Column AN (Col 40) onwards
    for s_idx, staff_name in enumerate(staff_list):
        staff_start_c = 40 + (s_idx * 14)
        staff_end_c = staff_start_c + 13
        
        ws_cons.merge_cells(start_row=1, start_column=staff_start_c, end_row=1, end_column=staff_end_c)
        r1_staff = ws_cons.cell(row=1, column=staff_start_c, value=staff_name)
        r1_staff.font = font_title
        r1_staff.fill = fill_green_header
        r1_staff.alignment = align_center
        
        for k_idx, (kpi_name, _) in enumerate(KPI_CATEGORIES):
            c_num = staff_start_c + k_idx
            
            h_cell = ws_cons.cell(row=2, column=c_num, value=kpi_name)
            h_cell.font = font_header
            h_cell.fill = fill_gray_header
            h_cell.alignment = align_center
            
            r3_cell = ws_cons.cell(row=3, column=c_num, value=0)
            r3_cell.font = font_bold
            r3_cell.fill = fill_gold_total
            r3_cell.alignment = align_center
            
            ws_cons.column_dimensions[get_column_letter(c_num)].width = 15
            
        for r in range(1, 4):
            for c in range(staff_start_c, staff_end_c + 1):
                if c == staff_end_c:
                    ws_cons.cell(row=r, column=c).border = cluster_divider_border
                else:
                    ws_cons.cell(row=r, column=c).border = thin_cell_border
                    
        for r in range(4, 54):
            for c in range(staff_start_c, staff_end_c + 1):
                if c == staff_end_c:
                    ws_cons.cell(row=r, column=c).border = cluster_divider_border
                else:
                    ws_cons.cell(row=r, column=c).border = thin_cell_border

    # -------------------------------------------------------------
    # 3. TABS 3 to 33: Daily Tabs ('1ST' to '31st')
    # -------------------------------------------------------------
    daily_headers = ["NAME", "DESIGNATION"] + [kpi[0] for kpi in KPI_CATEGORIES]
    
    for day_int in range(1, 32):
        tab_title = get_ordinal_tab_name(day_int)
        ws_day = wb.create_sheet(title=tab_title)
        ws_day.views.sheetView[0].showGridLines = True
        
        for c_idx, h in enumerate(daily_headers, start=1):
            cell = ws_day.cell(row=1, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_navy if c_idx <= 2 else fill_gray_header
            cell.alignment = align_center
            cell.border = header_border
            
        for s_idx, staff_name in enumerate(staff_list):
            fo_start_r = 2 + (s_idx * 40)
            fo_end_r = fo_start_r + 39
            
            cell_name = ws_day.cell(row=fo_start_r, column=1, value=staff_name)
            cell_name.font = font_bold
            cell_name.alignment = align_left
            
            cell_des = ws_day.cell(row=fo_start_r, column=2, value="Field Officer")
            cell_des.font = font_regular
            cell_des.alignment = align_center
            
            for r in range(fo_start_r, fo_end_r + 1):
                is_bottom_row = (r == fo_end_r)
                for c in range(1, len(daily_headers) + 1):
                    if is_bottom_row:
                        ws_day.cell(row=r, column=c).border = block_bottom_border
                    else:
                        ws_day.cell(row=r, column=c).border = thin_cell_border
                    ws_day.cell(row=r, column=c).alignment = align_center if c > 2 else (align_left if c == 1 else align_center)
                    
        ws_day.column_dimensions['A'].width = 22
        ws_day.column_dimensions['B'].width = 14
        for c in range(3, len(daily_headers) + 1):
            ws_day.column_dimensions[get_column_letter(c)].width = 15

    if default_sheet.title in wb.sheetnames and len(wb.sheetnames) > 33:
        wb.remove(default_sheet)
    elif "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[SUCCESS] Generated Template for {district}: {output_path} ({len(wb.sheetnames)} tabs)")

def main():
    print("="*60)
    print("Generating Master KPI Excel Templates with Premium Borders...")
    print("="*60)
    for district, staff in DISTRICT_STAFF.items():
        out_file = f"templates/template_{district}.xlsx"
        generate_district_template(district, staff, out_file)
    print("\n[ALL TEMPLATES GENERATED WITH PREMIUM BORDERS]")

if __name__ == "__main__":
    main()