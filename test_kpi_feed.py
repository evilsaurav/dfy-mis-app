import openpyxl
import re
from datetime import datetime

wb = openpyxl.load_workbook("templates/template_Jamui.xlsx")

print("Jamui Sheet Names:", wb.sheetnames[:5])

# Build case-insensitive sheet map
sheet_map = {name.strip().lower(): name for name in wb.sheetnames}
print("Sheet map keys for 1st/1st:", sheet_map.get("1st"), sheet_map.get("1st".lower()))

# Check performance sheet targets
if "Performance sheet" in wb.sheetnames:
    ws_p = wb["Performance sheet"]
    for r in range(2, 6):
        name = ws_p.cell(row=r, column=1).value
        print(f"Performance row {r}: Name='{name}', Target={ws_p.cell(row=r, column=3).value}")
